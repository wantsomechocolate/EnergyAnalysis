from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.style import Color, Fill
from openpyxl.cell import Cell
import numpy as np
import datetime
import pandas as pd
import xlsxwriter

#wb = load_workbook(book_name)

##wb = Workbook()
##
##ws=wb.create_sheet(0,"Test")
##
##c=ws.cell(row=0,column=0)
##
##c.value=5
##
##c.style.font.color.index=Color.GREEN

## To fit everything on one sheet,
## Put data streams accross and months down. 

input_date=datetime.datetime(2013,8,14)

def get_calander_from_date(date):

    #date=datetime.datetime(2013,2,14)
    fdom=datetime.datetime(date.year,date.month,1)
    fdow=fdom.isoweekday()

    fdonm=datetime.datetime(date.year, date.month+1,1)
    dim=(fdonm-fdom).days

    calander=np.zeros(6*7).reshape((6,7))


    for i in range(dim):
        index=i+fdow
        row=int(index/7)
        col=index%7
        calander[row][col]=i+1


    what=pd.DataFrame(calander,[1,2,3,4,5,6],columns=['sun','mon','tue','wen','thu','fri','sat'])

    return what



what=get_calander_from_date(input_date)
output_book = pd.ExcelWriter('Calander.xlsx')
what.to_excel(output_book,"June")
output_book.save()
output_book.close()




wb = load_workbook('Calander.xlsx')
ws=wb.get_sheet_by_name('June')

date=input_date

fdom=datetime.datetime(date.year,date.month,1)
fdow=fdom.isoweekday()

fdonm=datetime.datetime(date.year, date.month+1,1)
dim=(fdonm-fdom).days


for i in range(dim):
    index=i+fdow
    row=int(index/7)
    col=index%7
    c=ws.cell(row=row+1, column=col+1)
    c.style.font.color.index = Color.GREEN
	

wb.save('Calander.xlsx')






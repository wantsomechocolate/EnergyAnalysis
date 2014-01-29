##---------------------------------------------------------------------------------------------------------------------------------------------------

import os, time, datetime
import numpy as np, pylab as pl
import wam as wam, wamo as wamo
from dateutil import parser
from marbles import glass as chan
from openpyxl import Workbook, load_workbook

#import datetime
#from openpyxl import load_workbook
#import wamo as wamo
#import pylab as pl

##---------------------------------------------------------------------------------------------------------------------------------------------------

divider="\n---------------------------------------------------------------------------\n"

##-----------------------------------------------------------------------

print divider,"-------------------Welcome to Insiderate (In-sid-er-ate)-------------------",divider

##-----------------------------------------------------------------------

## Have user navigate to desired book and show them what they chose.
print "--Please navigate to the .xlsx file containing your data\n"
book_name=chan.getPath(os.getcwd())  #,ext_list=['.xlsx'])
print "--You chose to analyze"+book_name

##-----------------------------------------------------------------------

##Get range on which to get average day statistics from. 
print divider
print "--You will now enter the date range for the QUARTER or MONTH that this report is for\n"
single_day_stats_date_range=wam.get_date_range_from_user(False)

##-----------------------------------------------------------------------

## How many similar days do you want to return?
print divider+"\n--Now you have to tell me how many days to be used when calculating the band. \
For 1 year, put 3, for 1.5 years, put 4, for 2 or more years, put 5. 6 is max\n"
default_choice=5
num_matches=chan.getIntegerInput(3,6,"--Just press enter to use the number brackets ["+str(default_choice)+"]> ",default_choice,[])

##-----------------------------------------------------------------------
time_list=[]
time_list.append(time.time())
##-----------------------------------------------------------------------

##-----------------------------------------------------------------------

print divider+"\n--Pulling interval and weather data in from the workbook: ",
a_work_book=wamo.MyWorkbook(book_name)

##-----------------------------------------------------------------------

##-----------------------------------------------------------------------
time_list.append(time.time())
print str(int(time_list[-1]-time_list[-2]))+" seconds"
##-----------------------------------------------------------------------


#------------------------------------------------------------------------

print "--Preparing energy data for analysis: ",
interval_data_no_headings=a_work_book.work_book_data_no_headings["Interval Usage"]
interval_data_by_day=wam.interval2day(interval_data_no_headings)
interval_data_object=wamo.IntervalData(interval_data_no_headings, interval_data_by_day[0])
interval_data_object_steam=wamo.IntervalData(interval_data_no_headings, interval_data_by_day[1])
interval_time=interval_data_object.date_list
interval_usage_elec=interval_data_object.data_list
interval_usage_steam=interval_data_object_steam.data_list
interval_time_by_day_elec=interval_data_object.datetime_list_by_day
interval_usage_by_day_elec=interval_data_object.data_list_by_day
interval_time_by_day_steam=interval_data_object_steam.datetime_list_by_day
interval_usage_by_day_steam=interval_data_object_steam.data_list_by_day
#------------------------------------------------------------------------

#------------------------------------------------------------------------
time_list.append(time.time())
print str(int(time_list[-1]-time_list[-2]))+" seconds"
#------------------------------------------------------------------------

#------------------------------------------------------------------------

print "--Creating temperature data for analysis: ",
interval_weather_no_headings=a_work_book.work_book_data_no_headings["Interval Temp"]                              
interval_weather_by_day=wam.interval2day(interval_weather_no_headings)             
weather_object=wamo.IntervalData(interval_weather_no_headings,interval_weather_by_day[0])
time_stamp_np=weather_object.datetime_list
wbt_np=weather_object.data_list
date_list=weather_object.date_list
ts_by_day=weather_object.datetime_list_by_day
wbt_by_day=weather_object.data_list_by_day
wbt_daily_ave=weather_object.averages_by_day

#------------------------------------------------------------------------
time_list.append(time.time())
print str(int(time_list[-1]-time_list[-2]))+" seconds"
#------------------------------------------------------------------------

##----------------------------------------------------------------------------------------
## Line between old and new
##----------------------------------------------------------------------------------------

#------------------------------------------------------------------------

## This should pull from a seperately maintained file.
print divider+"\n--Getting the holidays to exclude from the analysis"
print "--The 2011-2014 holidays are currently hardcoded into the source"
print "--It is every Federal Holiday, plus the day after T-Gives"
holidays = wam.getholidays()

#------------------------------------------------------------------------

## Find the N most similar days to each day
print divider+"\n--Finding similar days based on daily average wetbulb temerature"
print "--We use wetbulb temperature because it more accurately reflects"
print "--Latent cooling requirements in the summer time. We use average"
print "--day because a better approach has not yet been implemented : ",
similar_days_by_day=wam.get_n_closest_matches_for_each_item_in_list(wbt_daily_ave, num_matches, date_list, holidays)

## make a list with the right dimensions
similar_days_by_DATE=[]
for i in range(len(similar_days_by_day)):
    similar_days_by_DATE.append([])

## Because the get_n_closest..... functions returns a list of list indices instead of a list of datetime objects
## Use those indicies to get the corresponding datetime objects.
for i in range(len(similar_days_by_day)):
    for j in range(len(similar_days_by_day[i])):
        similar_days_by_DATE[i].append(date_list[similar_days_by_day[i][j]])

## For printing purposes
similar_days_by_DATE_zipped = zip(*similar_days_by_DATE)

#------------------------------------------------------------------------
time_list.append(time.time())
print str(int(time_list[-1]-time_list[-2]))+" seconds"
#------------------------------------------------------------------------


##------------------------------------------------------------------------------
##---------------------Apply weather findings to interval data---------------------

print divider+"\n--Now that we have the similar days, we need to get the data"
print "--from each of those similar days on an interval basis"
print "--in order to be able to do the next step"
print "--So let's do that: ",


## Function name is pretty weak, you probably get what it's doing
similar_days_interval_usage_elec=wam.use_list_of_list_of_indices_to_group_a_list_of_lists(interval_usage_by_day_elec,similar_days_by_day)

## And one for steam two, these functions should just be able to handle different lists
similar_days_interval_usage_steam=wam.use_list_of_list_of_indices_to_group_a_list_of_lists(interval_usage_by_day_steam,similar_days_by_day)

##--------------------------------------------------------
time_list.append(time.time())
print str(round(time_list[-1]-time_list[-2],1))+" seconds"
##--------------------------------------------------------


##----------------------Do simple stat calcs------------------------------------

print divider+"\n--Reorganizing the data and calculating average and stdev for each interval value: ",

## Reorganizing the data to make it easier to get average days and calculate stdev and stuff
similar_days_interval_usage_by_interval_elec=wam.zip_all_items_of_a_list(similar_days_interval_usage_elec)

## And again for steeeeeaaaaam (lame)
similar_days_interval_usage_by_interval_steam=wam.zip_all_items_of_a_list(similar_days_interval_usage_steam)

## This function returns a list with three lists that are the same shape as orig list except the deepest list turned into a number
stats_elec=wam.get_ave_std_of_list_of_list_of_list(similar_days_interval_usage_by_interval_elec)

## Average
year_of_average_days_elec=stats_elec[0]

## Upper bound - Can be done in excel, don't bother
year_of_std_upper_elec=stats_elec[1]

## Lower bound - Can be done in excel, don't bother
year_of_std_lower_elec=stats_elec[2]

## Standard deviation
year_of_std_elec=stats_elec[3]

## And now for steam!
stats_steam=wam.get_ave_std_of_list_of_list_of_list(similar_days_interval_usage_by_interval_steam)
year_of_average_days_steam, year_of_std_upper_steam, year_std_lower_steam, year_of_std_steam = stats_steam


## Preparing some stuff for the graphs. 
interval_averages_elec=[]
interval_upper_bound_elec=[]
interval_lower_bound_elec=[]
interval_std_elec=[]
for i in range(len(year_of_average_days_elec)):
    for j in range(len(year_of_average_days_elec[i])):
        interval_averages_elec.append(year_of_average_days_elec[i][j])
        interval_upper_bound_elec.append(year_of_std_upper_elec[i][j])
        interval_lower_bound_elec.append(year_of_std_lower_elec[i][j])
        interval_std_elec.append(year_of_std_elec[i][j])

interval_averages_steam=[]
interval_upper_bound_steam=[]
interval_lower_bound_steam=[]
interval_std_steam=[]
for i in range(len(year_of_average_days_steam)):
    for j in range(len(year_of_average_days_steam[i])):
        interval_averages_steam.append(year_of_average_days_steam[i][j])
        interval_upper_bound_steam.append(year_of_std_upper_elec[i][j])
        interval_lower_bound_steam.append(year_of_std_lower_elec[i][j])
        interval_std_steam.append(year_of_std_steam[i][j])


##--------------------------------------------------------
time_list.append(time.time())
print str(round(time_list[-1]-time_list[-2],1))+" seconds"
##--------------------------------------------------------


print "--Getting the baseline for each day",
## All of these parameters were based on a single data set. This section of the program really needs some work. 
num_of_min_values=10
start_time=1*4
end_time=12*4 #12 noon - this way the baseline isn't thrown off by values happening at end of day

baseline_by_day_elec=wam.get_baseline_by_day(interval_usage_by_day_elec,num_of_min_values, start_time, end_time)

start_time=0
end_time=4*4
num_of_min_values=5

baseline_by_day_steam=wam.get_baseline_by_day(interval_usage_by_day_steam,num_of_min_values, start_time, end_time)


##--------------------------------------------------------
time_list.append(time.time())
print str(round(time_list[-1]-time_list[-2],1))+" seconds"
##--------------------------------------------------------


print "--Getting the startup time for each day",
percent_above_baseline=0.03
thresh=8

start_time_each_day_elec=wam.get_start_time_each_day(interval_time_by_day_elec,
                                                     interval_usage_by_day_elec,
                                                     baseline_by_day_elec,
                                                     percent_above_baseline,
                                                     thresh)

start_time_each_day_steam=wam.get_start_time_each_day(interval_time_by_day_steam,
                                                      interval_usage_by_day_steam,
                                                      baseline_by_day_steam,
                                                      percent_above_baseline,
                                                      thresh)


#------------------------------------------------------------------------
time_list.append(time.time())
print str(round(time_list[-1]-time_list[-2],1))+" seconds"
#------------------------------------------------------------------------


print "--Getting the shutdown time for each day",

thresh_end=1
percent_above_baseline=0.03

end_time_each_day_elec=wam.get_end_time_each_day(interval_time_by_day_elec,
                                                 interval_usage_by_day_elec,
                                                 baseline_by_day_elec,
                                                 start_time_each_day_elec,
                                                 percent_above_baseline,
                                                 thresh_end)

end_time_each_day_steam=wam.get_end_time_each_day(interval_time_by_day_steam,
                                                  interval_usage_by_day_steam,
                                                  baseline_by_day_steam,
                                                  start_time_each_day_steam,
                                                  percent_above_baseline,
                                                  thresh_end)

## The above works ok, maybe I should plot the previous days basline in orange or something to show more info.


#------------------------------------------------------------------------
time_list.append(time.time())
print str(round(time_list[-1]-time_list[-2],1))+" seconds"
#------------------------------------------------------------------------


#------------------------------------------------------------------------
print "--Getting the single day statistics based on the time period you entered earlier",

#------------------------------------------------------------------------
single_day_stats_elec=wam.get_stats_by_day_in_range(interval_usage_by_day_elec, date_list, single_day_stats_date_range)
wk_day_average_for_date_range_elec=single_day_stats_elec[0]
wk_end_average_for_date_range_elec=single_day_stats_elec[1]
peak_day_for_date_range_elec=single_day_stats_elec[2]
peak_date_for_date_range_elec=single_day_stats_elec[3]
#------------------------------------------------------------------------

#------------------------------------------------------------------------
single_day_stats_steam=wam.get_stats_by_day_in_range(interval_usage_by_day_steam, date_list, single_day_stats_date_range)
wk_day_average_for_date_range_steam=single_day_stats_steam[0]
wk_end_average_for_date_range_steam=single_day_stats_steam[1]
peak_day_for_date_range_steam=single_day_stats_steam[2]
peak_date_for_date_range_steam=single_day_stats_steam[3]
#------------------------------------------------------------------------

#------------------------------------------------------------------------
single_day_stats_wbt=wam.get_stats_by_day_in_range(wbt_by_day, date_list, single_day_stats_date_range)
wk_day_average_for_date_range_wbt=single_day_stats_wbt[0]
wk_end_average_for_date_range_wbt=single_day_stats_wbt[1]
peak_day_for_date_range_wbt=single_day_stats_wbt[2]
peak_date_for_date_range_wbt=single_day_stats_wbt[3]
#------------------------------------------------------------------------

#------------------------------------------------------------------------
start_time_for_plotting_average_day=datetime.datetime(2000,1,1,0,0)
time_range_for_plotting_average_day=[]
for i in range(96):
    time_range_for_plotting_average_day.append(start_time_for_plotting_average_day+datetime.timedelta(minutes=15*i))
#------------------------------------------------------------------------
    

#------------------------------------------------------------------------
ts_year_of_days=[]

#ts_by_day is a list of every day - each day containing a list of each hour - each hour containing a datetime object for that hour
# That means that for every iteration of the loop below, day_of_hours will contain a list of 24 datetime objects.
for day_of_hours in ts_by_day:
    try:
        #Try to append a single datetime object per day
        ts_year_of_days.append(day_of_hours[0])
        # and also add the day of the year to day_of_year - this should be changed, it doesn't work when crossing over from year to year
        #day_of_year.append(day_of_hours[0].timetuple()[7])
        
    except:
        # If the above fails, it most likely means that there was no data for the day in question, just append "err"
        ts_year_of_days.append("err")
        #day_of_year.append("err")

similar_days_by_day_zipped = zip(*similar_days_by_day)

ave_wbt_of_similar_days=[]

for i in range(num_matches):
    ave_wbt_of_similar_days.append([])

for i in range(len(similar_days_by_day_zipped)):
    for j in range(len(similar_days_by_day_zipped[i])):
        ave_wbt_of_similar_days[i].append(wbt_daily_ave[similar_days_by_day_zipped[i][j]])
#------------------------------------------------------------------------


#------------------------------------------------------------------------
time_list.append(time.time())
print str(round(time_list[-1]-time_list[-2],1))+" seconds"
#------------------------------------------------------------------------
print divider


bucket_date_range=wam.get_bucket_date_range_from_user(end_date=single_day_stats_date_range[1])
#bucket_date_range=single_day_stats_date_range[1]

try:
    start_date_index=date_list.index(bucket_date_range[0])
except ValueError:
    start_date_index=0

try:
    end_date_index=date_list.index(bucket_date_range[1])
except:
    end_date_index=-1
    print "--Something went wrong getting the end date"
    print "--Defaulting to the last date in list"


## Graph the average weekday so that the user can get the operating hours visually
## I calculate the start and stop times for each day, but it isn't very robust
## so this is currently the accepted method. 

## I show the electric graph here


print "--I'm showing you a graph now for electric usage. You have to observe the operating \
hours and then enter them (I would right them down). Remember that 'open' is the more common case.\n"

ave_day_plot=pl.plot_date(time_range_for_plotting_average_day,wk_day_average_for_date_range_elec,'g-')
ave_day_plot=pl.plot_date(time_range_for_plotting_average_day,wk_end_average_for_date_range_elec,'b-')
ave_day_plot=pl.plot_date(time_range_for_plotting_average_day,peak_day_for_date_range_elec,'r-')
pl.show()


## Then the user has to close the graph and enter those numbers. 
bucket_open_closed_elec=wam.get_operating_hours_from_user()

bucket_operating_hours_by_day_elec=[]
for i in range((bucket_date_range[1]-bucket_date_range[0]).days):
    bucket_operating_hours_by_day_elec.append(bucket_open_closed_elec)
    

#print divider
print ""


## Then plot for steam so they can enter that in - eventually, the code will
## behave properly for any number of columns but I keep putting that off for some reason
print "--Now do the same thing for steam. Remember the first thing you do after seeing the graph is typing 'open', or 'closed'"
ave_day_plot=pl.plot_date(time_range_for_plotting_average_day,wk_day_average_for_date_range_steam,'g-')
ave_day_plot=pl.plot_date(time_range_for_plotting_average_day,wk_end_average_for_date_range_steam,'b-')
ave_day_plot=pl.plot_date(time_range_for_plotting_average_day,peak_day_for_date_range_steam,'r-')
pl.show()

## User enters operating hours in here. 
bucket_open_closed_steam=wam.get_operating_hours_from_user()

bucket_operating_hours_by_day_steam=[]
for i in range((bucket_date_range[1]-bucket_date_range[0]).days):
    bucket_operating_hours_by_day_steam.append(bucket_open_closed_steam)



bucketed_usage_elec=wam.get_bucketed_usage(bucket_operating_hours_by_day_elec, date_list, start_date_index, end_date_index,
                       interval_usage_by_day_elec)

bucketed_usage_steam=wam.get_bucketed_usage(bucket_operating_hours_by_day_steam, date_list, start_date_index, end_date_index,
                       interval_usage_by_day_steam)




print "\n--Getting the operating hours from you took: ",
#------------------------------------------------------------------------
time_list.append(time.time())
print str(round(time_list[-1]-time_list[-2],1))+" seconds"
#------------------------------------------------------------------------


#------------------------------------------------------------------------
print "--The total runtime to this point was: "+str(round(time_list[-1]-time_list[0],1))+" seconds"
#------------------------------------------------------------------------


print divider
##-----------------------------------------------------------------------------------------------
##-----------------------------Printing to Excel------------------------------
##-----------------------------------------------------------------------------------------------
print "--Printing results to excel: ",
output_book=chan.add_to_filename(book_name," - Results - "+str(int(time_list[0])))
wb = Workbook()


## Printing the interval analysis results---------------------------------------------------------

ws1=wb.create_sheet(0,"Interval Analysis")

column_headings=["Time Stamp",
                 "Electric Usage(kWh)",
                 "Average Elec Usage(kWh)",
                 "STDEV Elec(kWh)",
                 #"Ave+STD Elec (kWh)",
                 #"Ave-STD Elec (kWh)",
                 "Steam Usage (lbs)",
                 "Average Steam Usage (lbs)",
                 "STDEV Steam (lbs)",
                 #"Ave+STD Steam (lbs)",
                 #"Ave-STD Steam (lbs)",
                 ]

output_list=[interval_data_no_headings[0],
                 interval_usage_elec,
                 interval_averages_elec,
                 interval_std_elec,
                 #interval_upper_bound_elec,
                 #interval_lower_bound_elec,
                 interval_usage_steam,
                 interval_averages_steam,
                 interval_std_steam,
                 #interval_upper_bound_steam,
                 #interval_lower_bound_steam,
                 ]

## for all headings i
for i in range(len(column_headings)):
    c=ws1.cell(row=0,column=i)
    c.value=column_headings[i]

    ## for all rows j+1
    for j in range(len(output_list[i])):
        c=ws1.cell(row=j+1,column=i)
        c.value=output_list[i][j]


#--------------------------------------------------------------------------------------------
## Printing the daily analysis results

#Create Tab
ws2=wb.create_sheet(-1,"Similar Day Analysis")

#Generate Headings
day_anal_headings=["Day"]

for i in range(num_matches):
    day_anal_headings.append("Sim Day "+str(i+1))

day_anal_headings.append("Ave Wetbulb Temp")

for i in range(num_matches):
    day_anal_headings.append("Sim Day "+str(i+1)+"Ave Wetbulb")

#Generate Output List
output_list_by_day=[ts_year_of_days]

for i in range(num_matches):
    output_list_by_day.append(similar_days_by_DATE_zipped[i])

output_list_by_day.append(wbt_daily_ave)

for i in range(num_matches):
    output_list_by_day.append(ave_wbt_of_similar_days[i])

##Print to file
for i in range(len(day_anal_headings)):
    c2=ws2.cell(row=0,column=i)
    c2.value=day_anal_headings[i]

    for j in range(len(output_list_by_day[i])):
        c2=ws2.cell(row=j+1,column=i)
        c2.value=output_list_by_day[i][j]


#-----------------------------------------------------------------------------------------------------

#Create Tab
ws_oper=wb.create_sheet(-1,"Operating Hours")

#Create Headings
operating_hours_headings=["Day",
                          "Ave Wetbulb Temp",
                          "Start Time Elec",
                          "Stop Time Elec",
                          "Baseline Elec",
                          "Start Time Steam",
                          "Stop Time Steam",
                          "Baseline Steam"]


#Some prelin
baseline_by_day_to_print_elec=[]
for i in range(len(baseline_by_day_elec)):
    baseline_by_day_to_print_elec.append(baseline_by_day_elec[i][0])

baseline_by_day_to_print_steam=[]
for i in range(len(baseline_by_day_steam)):
    baseline_by_day_to_print_steam.append(baseline_by_day_steam[i][0])

#Put together data to print
operating_hours_data=[ts_year_of_days,
                      wbt_daily_ave,
                      start_time_each_day_elec,
                      end_time_each_day_elec,
                      baseline_by_day_to_print_elec,
                      start_time_each_day_steam,
                      end_time_each_day_steam,
                      baseline_by_day_to_print_steam]

#Print the data to file
for i in range(len(operating_hours_headings)):
    c2=ws_oper.cell(row=0,column=i)
    c2.value=operating_hours_headings[i]

    for j in range(len(operating_hours_data[i])):
        c2=ws_oper.cell(row=j+1,column=i)
        c2.value=operating_hours_data[i][j]


##-----------------------------------------------------------------------------------
## Printing the single day stat results

#Make Sheet
ws3=wb.create_sheet(-1,"Single Day Stats")

#Headings
single_day_stat_headings=["Average WkDay Elec",
                          "Average WkEnd Elec",
                          "Peak Day Elec",
                          "Average WkDay Steam",
                          "Average WkEnd Steam",
                          "Peak Day Steam"]

#Data
single_day_stat_data=[wk_day_average_for_date_range_elec,
                      wk_end_average_for_date_range_elec,
                      peak_day_for_date_range_elec,
                      wk_day_average_for_date_range_steam,
                      wk_end_average_for_date_range_steam,
                      peak_day_for_date_range_steam]

#Print
for i in range(len(single_day_stat_headings)):
    c3=ws3.cell(row=0,column=i)
    c3.value=single_day_stat_headings[i]

    for j in range(len(single_day_stat_data[i])):
        c3=ws3.cell(row=j+1,column=i)
        c3.value=single_day_stat_data[i][j]

#For this Tab I need to print some extra stuff. 
c3=ws3.cell(row=0, column=(len(single_day_stat_data)))
c3.value="Peak Date Elec"

c3=ws3.cell(row=1, column=(len(single_day_stat_data)))
c3.value=peak_date_for_date_range_elec

c3=ws3.cell(row=0, column=(len(single_day_stat_data)+1))
c3.value="Peak Date Steam"

c3=ws3.cell(row=1, column=(len(single_day_stat_data)+1))
c3.value=peak_date_for_date_range_steam




##-----------------------------------------------------------------------------------
## Printing the single day weather results

#Make Sheet
ws3=wb.create_sheet(-1,"Single Day Stats WBT")

#Headings
single_day_stat_WBT_headings=["Average WkDay WBT",
                              "Average WkEnd WBT",
                              "Peak Day WBT"]

#Data
single_day_stat_WBT_data=[wk_day_average_for_date_range_wbt,
                          wk_end_average_for_date_range_wbt,
                          peak_day_for_date_range_wbt]

#Print
for i in range(len(single_day_stat_WBT_headings)):
    c3=ws3.cell(row=0,column=i)
    c3.value=single_day_stat_WBT_headings[i]

    for j in range(len(single_day_stat_WBT_data[i])):
        c3=ws3.cell(row=j+1,column=i)
        c3.value=single_day_stat_WBT_data[i][j]


#For this Tab I need to print some extra stuff. 
c3=ws3.cell(row=0, column=(len(single_day_stat_WBT_data)))
c3.value="Peak Date WBT"

c3=ws3.cell(row=1, column=(len(single_day_stat_WBT_data)))
c3.value=peak_date_for_date_range_wbt



#---------------------------------------------------------------
## Printing bucketed usage

ws_buckets=wb.create_sheet(-1,"Bucketed Usage Elec")

bucket_headings=["Date",
                 "Usage Open Hours Elec",
                 "Usage Closed Hours Elec",
                 "Usage Open Hours Steam",
                 "Usage Closed Hours Steam"]

bucket_data=[bucketed_usage_elec[2],
             bucketed_usage_elec[0],
             bucketed_usage_elec[1],
             bucketed_usage_steam[0],
             bucketed_usage_steam[1]]

for i in range(len(bucket_headings)):
    c=ws_buckets.cell(row=0,column=i)
    c.value=bucket_headings[i]

    for j in range(len(bucket_data[i])):
        c=ws_buckets.cell(row=j+1,column=i)
        c.value=bucket_data[i][j]

##------------------------------------------------------------------------------------------------

wb.save(output_book)

##--------------------------------------------------------
time_list.append(time.time())
print str(round(time_list[-1]-time_list[-2],1))+" seconds"
##--------------------------------------------------------

####------------------------------------------------------
#### Experimenting with printing with pandas
##import pandas as pd
###df = pd.DataFrame(bucketed_usage_elec, columns=headers)
##df = pd.DataFrame(bucketed_usage_elec)
##df.to_excel('test.xlsx',"Bucketed Usage")
####------------------------------------------------------


print divider

raw_input("--Press 'ENTER' to exit")

print "--Exited Program"

##------------------------------------------------------------------------------------------------------------------------------------------


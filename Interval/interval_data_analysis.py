from openpyxl import Workbook
from openpyxl import load_workbook

import os

import numpy as np

#import scipy as sp

import pylab as pl

working_directory="C:/Users/James McGlynn/My Programs/Python Programs/Utility Data Analysis/Interval Data/"

os.chdir(working_directory)

book_name="interval_data_2012.xlsx"

book_path=working_directory+book_name

wb=load_workbook(book_path)

ws = wb.get_active_sheet()

last_occ_row=ws.rows[-1][0].row

#col = 0
time_stamp_list=[]
kwh_list=[]

for i in range(last_occ_row):
    c0=ws.cell(row=i, column=0)
    time_stamp_list.append(c0.value)

    c1=ws.cell(row=i, column=1)
    kwh_list.append(c1.value)

time_stamp_np=np.array(time_stamp_list[1:])
kwh_np=np.array(kwh_list[1:])

#kwh_plot=pl.plot(time_stamp_np,kwh_np)
kwh_plot=pl.plot_date(time_stamp_np,kwh_np)
#kwh_plot.setp(lines,'color','r','linewidth',2.0)

pl.show()








import datetime
import os, wam, time
import numpy as np
#import pylab as pl
from marbles import glass as chan


from openpyxl import Workbook
from openpyxl import load_workbook


class MyWorkbook(object):

    ## As soon as this class is called...
    def __init__(self, book_name):

        ## Load the workbook
        self.wb = load_workbook(book_name)

        ## Get all the sheet names - the order of this list will be the
        ## same order as the sheets in the book I believe (hope)
        self.sheet_names=self.wb.get_sheet_names()

        self.sheet_objects=[]

        #self.sheet_data_objects=[]
        
        ## For each sheet, go in and get the addresses and actual data for each row and
        ## and column that contain any data. 
        for i in range(len(self.sheet_names)):

            ## This initializes a sheet object, which takes the sheet name
            ## and returns the address of the smallest rectangle that includes
            ## all the data.
            self.sheet_objects.append(MySheet(self.wb, self.sheet_names[i]))

            ## This initializes a sheet data object which takes the addresses and
            ## returns the data.
            #self.sheet_data_objects.append(MySheetData(self.sheet_objects[i].sheet_data_range))

        ## This creates as many objects as sheets with the actual data inside (including the headings)
        self.work_book_data={}
        self.work_book_data_no_headings={}
        for i in range(len(self.sheet_names)):
            self.work_book_data[self.sheet_names[i]]=self.sheet_objects[i].sheet_data

            no_heading_intermediate=[]
            for k in range(len(self.sheet_objects[i].sheet_data)):
                no_heading_intermediate.append(self.sheet_objects[i].sheet_data[k][1:])

            self.work_book_data_no_headings[self.sheet_names[i]]=no_heading_intermediate


####        self.work_book_data_no_headings={}
####        for i in range(len(self.sheet_names)):
####            
####            interval_data_no_headings_inter=[]
####
####            for i in range(len(interval_data)):
####                interval_data_no_headings.append(interval_data[i][1:])
            

    def get_num_sheets(self):

        return len(self.sheet_objects)


class MySheet(object):

    def __init__(self, workbook, sheet_name):

        self.sheet=workbook.get_sheet_by_name(sheet_name)

        self.sheet_data_range=self.sheet.columns

        self.sheet_data=[]
        for i in range(len(self.sheet_data_range)):
            self.sheet_data.append([])

        for i in range(len(self.sheet_data_range)):
            for j in range(len(self.sheet_data_range[i])):
                self.sheet_data[i].append(self.sheet_data_range[i][j].value)

class IntervalData(object):

    def __init__(self, interval_data_def, interval_data_by_day_def):

        self.datetime_list=interval_data_def[0]
        
        self.data_list=interval_data_def[1]

        self.datetime_list_by_day=interval_data_by_day_def[0]

        self.data_list_by_day=interval_data_by_day_def[1]

        self.averages_by_day=self.list_of_lists_2_list_of_ave(self.data_list_by_day)


    def get_elapsed_days(self):
        
        return self.datetime_list_by_day[-1][0]-self.datetime_list_by_day[0][0]


    def list_of_lists_2_list_of_ave(self, list_of_lists_def):

        list_of_aves_def=[]
        
        for i in range(len(list_of_lists_def)):
            try:
                daily_ave_def=float(sum(list_of_lists_def[i]))/len(list_of_lists_def[i])
            except:
                daily_ave_def="err"
            list_of_aves_def.append(daily_ave_def)
        return list_of_aves_def

##    def get_average_weekday(self):
##    def get_average_weekend(self):
##    def get_peak_day(self):
##    def get_start_and_end_time_from_ave_wkday(self):
##    def get_bucketed_usage(self):
    

def interval2day(interval_data_def):


    number_of_non_date_columns_def=len(interval_data_def[1:])
    
    number_of_columns_def=len(interval_data_def)

    ## The datetime is assumed to be the first list 
    datetime_list_def=interval_data_def[0]

    ## Create space for the lists of data (exclude the date)
    data_lists_def=[]

    ## Now I have a list of at least one other list, but possibly more
    for i in range(1,number_of_columns_def):
        
        data_lists_def.append(interval_data_def[i])


    ## Get the first date (assumed to be the earliest date)
    current_date_def=datetime.datetime(datetime_list_def[0].year, datetime_list_def[0].month, datetime_list_def[0].day)

    ## Get the last date (assumed to be the most recent date)
    end_date_def=datetime.datetime(datetime_list_def[-1].year, datetime_list_def[-1].month, datetime_list_def[-1].day)

    ## Prepare for creation of date list
    date_list_def=[]
    
    ## This is done this way at the moment in case there are missing dates, at least every day will still have
    ## a space allocated for it. 
    while current_date_def<=end_date_def:
        date_list_def.append(current_date_def)
        current_date_def=current_date_def+datetime.timedelta(days=1)

    ## Make a number of unique lists
    unique_lists=[]
    for i in range(number_of_columns_def):
        unique_lists.append([])
        for j in range(len(date_list_def)):
            unique_lists[i].append([])

    datetime_list_by_day_def=unique_lists[0]
                                   
    data_lists_by_day_def=[]
    ## For as many columns of data there are
    for i in range(1,number_of_columns_def):
        ## Make room for that column of data to be sorted by day
        data_lists_by_day_def.append(unique_lists[i])

        
    ## Go through the huge list and put everything where it goes.
    ## FOR EVERY SINGLE DATA POINT in the original datetime list
    for i in range(len(datetime_list_def)):

        ## Strip the time off of the datetime in the interval datetime list
        interval_data_day_def=datetime.datetime(datetime_list_def[i].year,datetime_list_def[i].month,datetime_list_def[i].day)

        ## Then find the index for that day in the date list
        index_def=date_list_def.index(interval_data_day_def)
        #print index_def

        datetime_list_by_day_def[index_def].append(datetime_list_def[i])

        for k in range(len(interval_data_def[1:])):
            data_lists_by_day_def[k][index_def].append(data_lists_def[k][i])
                                       
    return_list_def=[]

    for i in range(len(interval_data_def[1:])):
        return_list_def.append([datetime_list_by_day_def,data_lists_by_day_def[i]])

    return return_list_def


book_name='C:/Users/James McGlynn/Documents/GitHub/EnergyAnalysis/Tool/ZZ - Test.xlsx'
#book_name=chan.getPath(os.getcwd())

time_list=[]
time_list.append(time.time())
print "Initializing workbook: ",


a_work_book=MyWorkbook(book_name)


time_list.append(time.time())
print str(int(time_list[-1]-time_list[-2]))+" seconds"


print "breaking up interval data by day: ",
interval_data_no_headings=a_work_book.work_book_data_no_headings["Interval Usage"]
interval_data_by_day=interval2day(interval_data_no_headings)
interval_data_object=IntervalData(interval_data_no_headings, interval_data_by_day[0])


time_list.append(time.time())
print str(int(time_list[-1]-time_list[-2]))+" seconds"


print "breaking up weather data by day: ",
interval_weather_no_headings=a_work_book.work_book_data_no_headings["Interval Temp"]                              
interval_weather_by_day=interval2day(interval_weather_no_headings)             
weather_object=IntervalData(interval_weather_no_headings,interval_weather_by_day[0])


time_list.append(time.time())
print str(int(time_list[-1]-time_list[-2]))+" seconds"

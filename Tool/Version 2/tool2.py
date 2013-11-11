## Interval Data Analysis Tool - Written by James McGlynn on his own time, on his own computer, while working
## for CodeGreen Solutions. Circa 2013 (Jun - November)

#dont make date list depend on minute value of time stamp, always reset it to zero



## This analysis should really use a minimum of two years of data.

## Future Considerations are:
## 1.)  Better automate creation of spreadsheet for use by tool
## 2.)  Fix Graphing Utility for easy exploration of results
## 3.)  Retrive holidays automatically and remove them from analysis of similar weather days.
## 4.)  Enhance the comparison of weather conditions - use more prior days (2 or three at most)
##          and try to compare at every point instead of using a daily average.
## 5.)  Create an exe to run using excel macro that will do the analysis all in excel.
## 6.)  Make standard deviation directional
## 7.)  Make number of similar days to be used in analysis dependant upon number of total days

## The operating hours part still needs serious work!

## -------------------IMPORTS--------------------

import datetime, os, wam, time
import numpy as np
#import pylab as pl
from marbles import glass as chan

from openpyxl import Workbook
from openpyxl import load_workbook

## ------------------Let's Begin-----------------

## Returns 2012 and 2013 Federal Holidays as a list of datetime objects.
holidays = wam.getholidays()

## Ask user to navigate to file
print "Select file using askopenfile dialog."
book_name=chan.getPath(os.getcwd())

## Ask user for date range to analyze
print "What date range would you like to use for the single day stats?"
single_day_stats_date_range=wam.get_date_range_from_user(False)

## How many similar days do you want to return? Three per year of data might be ok.
print "Please select the number of similar days you want to find for each day: ",
num_matches_raw=raw_input("> ")
num_matches=5

## (Time keeping stuff - not important)
##------------------------------------------------
time_list=[]
time_list.append(time.time())
##------------------------------------------------


## ---------------------------Get data from spreadsheet-----------------------------

print "Retrieving weather data from "+book_name+": ",

## Unfortunately I have to rely on this being true :(
weather_sheet_name='Interval Temp'
## This shouldn't be necessary
number_of_columns=2

## This function returns a list of two lists
tstamp_wbulb=wam.xlsx2np(book_name, weather_sheet_name, number_of_columns)

## The first is hourly time stamp - I checked and these things come out in order form oldest to newest
time_stamp_np=tstamp_wbulb[0]

## The second is hourly wetbulb temp
wbt_np=tstamp_wbulb[1]


## Use the data from the spreadsheet to define a list that has the date starting from start date to end date
## In case there are gaps. At the moment the data has to be in date order oldest to newest. 
current_date=time_stamp_np[0]
end_date=time_stamp_np[-1]
date_list=[] # - I checked and this is also in order from oldest to newest
while current_date<=end_date:
    date_list.append(current_date)
    current_date=current_date+datetime.timedelta(days=1)

#----------------------------------------------------------
time_list.append(time.time())
print str(round(time_list[-1]-time_list[-2],1))+" seconds"
#----------------------------------------------------------

## ------------------Get the average wet bulb temperature for each day---------------
print "Finding the average wetbulb temperature for each day: ",

#number_of_columns=3

## This function also returns a list of two lists
by_day=wam.interval2day(time_stamp_np,wbt_np)

## One list of the time stamp BY DAY 
ts_by_day=by_day[0] #----this now comes out in order

## One list of the wetbulb temp BY DAY
wbt_by_day=by_day[1]

## This function takes a list of lists and returns a list of averages
wbt_daily_ave=wam.list_of_lists_2_list_of_ave(wbt_by_day)

#-----------------------------------------------------------
time_list.append(time.time())
print str(round(time_list[-1]-time_list[-2],1))+" seconds"
#-----------------------------------------------------------

## ---------------------Find the N most similar days to each day--------------------
print "Finding similar days based on a single criteria (Average WBT): ",

## This function takes a list of numbers, the number of closest matches to get, and a date list
## The date list is used as a criteria because I actually only want to consider
## Similar WB temps that occur on the same day of the week.

similar_days_by_day=wam.get_n_closest_matches_for_each_item_in_list(wbt_daily_ave, num_matches, date_list, holidays)

## make a list with the right dimensions
similar_days_by_DATE=[]
for i in range(len(similar_days_by_day)):
    similar_days_by_DATE.append([])

## Becase the get_n_closest..... functions returns a list of list indices instead of a list of datetime objects
## Use those indicies to get the corresponding datetime objects.
for i in range(len(similar_days_by_day)):
    for j in range(len(similar_days_by_day[i])):
        similar_days_by_DATE[i].append(date_list[similar_days_by_day[i][j]])

## For printing purposes
similar_days_by_DATE_zipped = zip(*similar_days_by_DATE)



#----------------------------------------------------------
time_list.append(time.time())
print str(round(time_list[-1]-time_list[-2],1))+" seconds"
#----------------------------------------------------------

## ---------------------Get interval usage data---------------------

print "Retrieving interval data from spreadsheet: ",

## Must be true
interval_data_sheet_name='Interval Usage'

## Shouldn't be necessary
number_of_columns=3

## This function returns a list of lists
interval_data=wam.xlsx2np(book_name, interval_data_sheet_name,number_of_columns)

## The first is interval time stamp
interval_time=interval_data[0]

## The second is interval data (in this case electric)
interval_usage_elec=interval_data[1]

## The third if it is there is probably steam
interval_usage_steam=interval_data[2]

#---------------------------------------------------------------
time_list.append(time.time())
print str(round(time_list[-1]-time_list[-2],1))+" seconds"
#---------------------------------------------------------------


print "Breaking up interval data by day: ",

## This function also returns a list of two lists, I have to call it twice right now because
## I'm assuming there will always be exactly two streams of data (STUPID!)
interval_by_day_elec=wam.interval2day(interval_time,interval_usage_elec)

## Calling it again for steam - pretty redundant, I should fix this
interval_by_day_steam=wam.interval2day(interval_time,interval_usage_steam)

## One list of the time stamp BY DAY 
interval_time_by_day_elec=interval_by_day_elec[0]

## One list of the elec usage BY DAY
interval_usage_by_day_elec=interval_by_day_elec[1]

## A list of the time stamp for steam (Will always be identical to the electric one)
interval_time_by_day_steam=interval_by_day_steam[0]

## A list of the steam usage BY DAY
interval_usage_by_day_steam=interval_by_day_steam[1]

#-------------------------------------------------------------
time_list.append(time.time())
print str(round(time_list[-1]-time_list[-2],1))+" seconds"
#-------------------------------------------------------------


##---------------------Apply weather findings to interval data---------------------

print "Gathering interval usage for days that had similar weather: ",

## Function name is pretty weak, you probably get what it's doing
similar_days_interval_usage_elec=wam.use_list_of_list_of_indices_to_group_a_list_of_lists(interval_usage_by_day_elec,similar_days_by_day)

## And one for steam two, these functions should just be able to handle different lists
similar_days_interval_usage_steam=wam.use_list_of_list_of_indices_to_group_a_list_of_lists(interval_usage_by_day_steam,similar_days_by_day)

##--------------------------------------------------------
time_list.append(time.time())
print str(round(time_list[-1]-time_list[-2],1))+" seconds"
##--------------------------------------------------------

##----------------------Do simple stat calcs------------------------------------

print "Reorganizing the data and calculating average and stdev for each interval value: ",

## Reorganizing the data to make it easier to get average days and calculate stdev and stuff
similar_days_interval_usage_by_interval_elec=wam.zip_all_items_of_a_list(similar_days_interval_usage_elec)

## And again for steeeeeaaaaam (lame)
similar_days_interval_usage_by_interval_steam=wam.zip_all_items_of_a_list(similar_days_interval_usage_steam)

## This function returns a list with three lists that are the same shape as orig list except the deepest list turned into a number
stats_elec=wam.get_ave_std_of_list_of_list_of_list(similar_days_interval_usage_by_interval_elec)

## Average
year_of_average_days_elec=stats_elec[0]

## Upper bound - Can be done in excel, don't bother
year_of_std_upper_elec=stats_elec[1]

## Lower bound - Can be done in excel, don't bother
year_of_std_lower_elec=stats_elec[2]

## Standard deviation
year_of_std_elec=stats_elec[3]

## And now for steam!
stats_steam=wam.get_ave_std_of_list_of_list_of_list(similar_days_interval_usage_by_interval_steam)

######The above line is getting an error use this code to check for the cause
######for i in range(len(similar_days_interval_usage_by_interval_steam)):
######	for j in range(len(similar_days_interval_usage_by_interval_steam[i])):
######		if None in similar_days_interval_usage_by_interval_steam[i][j]:
######			print i,j

########for item in sample_list_np:
########	try:
########		new_list.append(float(item))
########	except:
########		pass

year_of_average_days_steam=stats_steam[0]
year_of_std_upper_steam=stats_steam[1]
year_of_std_lower_steam=stats_steam[2]
year_of_std_steam=stats_steam[3]

interval_averages_elec=[]
interval_upper_bound_elec=[]
interval_lower_bound_elec=[]
interval_std_elec=[]
for i in range(len(year_of_average_days_elec)):
    for j in range(len(year_of_average_days_elec[i])):
        interval_averages_elec.append(year_of_average_days_elec[i][j])
        interval_upper_bound_elec.append(year_of_std_upper_elec[i][j])
        interval_lower_bound_elec.append(year_of_std_lower_elec[i][j])
        interval_std_elec.append(year_of_std_elec[i][j])

interval_averages_steam=[]
interval_upper_bound_steam=[]
interval_lower_bound_steam=[]
interval_std_steam=[]
for i in range(len(year_of_average_days_steam)):
    for j in range(len(year_of_average_days_steam[i])):
        interval_averages_steam.append(year_of_average_days_steam[i][j])
        interval_upper_bound_steam.append(year_of_std_upper_elec[i][j])
        interval_lower_bound_steam.append(year_of_std_lower_elec[i][j])
        interval_std_steam.append(year_of_std_steam[i][j])


##--------------------------------------------------------
time_list.append(time.time())
print str(round(time_list[-1]-time_list[-2],1))+" seconds"
##--------------------------------------------------------

print "Getting the baseline for each day",

## get baseline by day Or should it be for the morning?

## All of these parameters were based on a single data set. This section of the program really needs some work. 
num_of_min_values=10
start_time=1*4
end_time=12*4 #12 noon - this way the baseline isn't thrown off by values happening at end of day

baseline_by_day_elec=wam.get_baseline_by_day(interval_usage_by_day_elec,num_of_min_values, start_time, end_time)

start_time=0
end_time=4*4
num_of_min_values=5

baseline_by_day_steam=wam.get_baseline_by_day(interval_usage_by_day_steam,num_of_min_values, start_time, end_time)

##--------------------------------------------------------
time_list.append(time.time())
print str(round(time_list[-1]-time_list[-2],1))+" seconds"
##--------------------------------------------------------

print "Getting the startup time for each day",

percent_above_baseline=0.03

thresh=8

start_time_each_day_elec=wam.get_start_time_each_day(interval_time_by_day_elec, interval_usage_by_day_elec, baseline_by_day_elec, percent_above_baseline, thresh)

start_time_each_day_steam=wam.get_start_time_each_day(interval_time_by_day_steam, interval_usage_by_day_steam, baseline_by_day_steam, percent_above_baseline, thresh)

##--------------------------------------------------------
time_list.append(time.time())
print str(round(time_list[-1]-time_list[-2],1))+" seconds"
##--------------------------------------------------------

print "Getting the shutdown time for each day",

thresh_end=1

percent_above_baseline=0.03

end_time_each_day_elec=wam.get_end_time_each_day(interval_time_by_day_elec, interval_usage_by_day_elec, baseline_by_day_elec, start_time_each_day_elec, percent_above_baseline, thresh_end)

end_time_each_day_steam=wam.get_end_time_each_day(interval_time_by_day_steam, interval_usage_by_day_steam, baseline_by_day_steam, start_time_each_day_steam, percent_above_baseline, thresh_end)

## The above works ok, maybe I should plot the previous days basline in orange or something to show more info.

##--------------------------------------------------------
time_list.append(time.time())
print str(round(time_list[-1]-time_list[-2],1))+" seconds"
##--------------------------------------------------------

##--------------------------------------------------------
print "The total runtime to this point was: "+str(round(time_list[-1]-time_list[0],1))+" seconds"
##--------------------------------------------------------

##--------------------------------------------------------
time_list.append(time.time())
print str(round(time_list[-1]-time_list[-2],1))+" seconds"
##--------------------------------------------------------

print "Getting the single day statistics using the date range you gave earlier!",
#single_day_stats_date_range=wam.get_date_range_from_user(False)

single_day_stats_elec=wam.get_stats_by_day_in_range(interval_usage_by_day_elec, date_list, single_day_stats_date_range)
wk_day_average_for_date_range_elec=single_day_stats_elec[0]
wk_end_average_for_date_range_elec=single_day_stats_elec[1]
peak_day_for_date_range_elec=single_day_stats_elec[2]
peak_date_for_date_range_elec=single_day_stats_elec[3]

single_day_stats_steam=wam.get_stats_by_day_in_range(interval_usage_by_day_steam, date_list, single_day_stats_date_range)
wk_day_average_for_date_range_steam=single_day_stats_steam[0]
wk_end_average_for_date_range_steam=single_day_stats_steam[1]
peak_day_for_date_range_steam=single_day_stats_steam[2]
peak_date_for_date_range_steam=single_day_stats_steam[3]

start_time_for_plotting_average_day=datetime.datetime(2000,1,1,0,0)
time_range_for_plotting_average_day=[]
for i in range(96):
    time_range_for_plotting_average_day.append(start_time_for_plotting_average_day+datetime.timedelta(minutes=15*i))
    



ts_year_of_days=[]

#ts_by_day is a list of every day - each day containing a list of each hour - each hour containing a datetime object for that hour
# That means that for every iteration of the loop below, day_of_hours will contain a list of 24 datetime objects.
for day_of_hours in ts_by_day:
    try:
        #Try to append a single datetime object per day
        ts_year_of_days.append(day_of_hours[0])
        # and also add the day of the year to day_of_year - this should be changed, it doesn't work when crossing over from year to year
        #day_of_year.append(day_of_hours[0].timetuple()[7])
        
    except:
        # If the above fails, it most likely means that there was no data for the day in question, just append "err"
        ts_year_of_days.append("err")
        #day_of_year.append("err")


similar_days_by_day_zipped = zip(*similar_days_by_day)

ave_wbt_of_similar_days=[]

for i in range(num_matches):
    ave_wbt_of_similar_days.append([])

for i in range(len(similar_days_by_day_zipped)):
    for j in range(len(similar_days_by_day_zipped[i])):
        ave_wbt_of_similar_days[i].append(wbt_daily_ave[similar_days_by_day_zipped[i][j]])




##--------------------------------------------------------
time_list.append(time.time())
print str(round(time_list[-1]-time_list[-2],1))+" seconds"
##--------------------------------------------------------



##-----------------------------Printing Shit to Excel------------------------------

print "Printing initial results to excel.",
output_book=chan.add_to_filename(book_name," - Results - "+str(int(time_list[0])))
wb = Workbook()

##-------------------------------------------------------------------------------
## time_list.append(time.time())
##-------------------------------------------------------------------------------

## Printing the interval analysis results---------------------------------------------------------

ws1=wb.create_sheet(0,"Interval Analysis")

column_headings=["Time Stamp",
                 "Electric Usage(kWh)",
                 "Average Elec Usage(kWh)",
                 "STDEV Elec(kWh)",
                 #"Ave+STD Elec (kWh)",
                 #"Ave-STD Elec (kWh)",
                 "Steam Usage (lbs)",
                 "Average Steam Usage (lbs)",
                 "STDEV Steam (lbs)",
                 #"Ave+STD Steam (lbs)",
                 #"Ave-STD Steam (lbs)"
                 ]

output_list=[interval_time,
             interval_usage_elec,
             interval_averages_elec,
             interval_std_elec,
             #interval_upper_bound_elec,
             #interval_lower_bound_elec,
             interval_usage_steam,
             interval_averages_steam,
             interval_std_steam,
             #interval_upper_bound_steam,
             #interval_lower_bound_steam
             ]

## for all headings i
for i in range(len(column_headings)):
    c=ws1.cell(row=0,column=i)
    c.value=column_headings[i]

    ## for all rows j+1
    for j in range(len(output_list[i])):
        c=ws1.cell(row=j+1,column=i)
        c.value=output_list[i][j]



#--------------------------------------------------------------------------------------------
## Printing the daily analysis results

#Create Tab
ws2=wb.create_sheet(-1,"Similar Day Analysis")

#Generate Headings
day_anal_headings=["Day"]

for i in range(num_matches):
    day_anal_headings.append("Sim Day "+str(i+1))

day_anal_headings.append("Ave Wetbulb Temp")

for i in range(num_matches):
    day_anal_headings.append("Sim Day "+str(i+1)+"Ave Wetbulb")

#Generate Output List
output_list_by_day=[ts_year_of_days]

for i in range(num_matches):
    output_list_by_day.append(similar_days_by_DATE_zipped[i])

output_list_by_day.append(wbt_daily_ave)

for i in range(num_matches):
    output_list_by_day.append(ave_wbt_of_similar_days[i])

##Print to file
for i in range(len(day_anal_headings)):
    c2=ws2.cell(row=0,column=i)
    c2.value=day_anal_headings[i]

    for j in range(len(output_list_by_day[i])):
        c2=ws2.cell(row=j+1,column=i)
        c2.value=output_list_by_day[i][j]


#-----------------------------------------------------------------------------------------------------

#Create Tab
ws_oper=wb.create_sheet(-1,"Operating Hours")

#Create Headings
operating_hours_headings=["Day",
                          "Ave Wetbulb Temp",
                          "Start Time Elec",
                          "Stop Time Elec",
                          "Baseline Elec",
                          "Start Time Steam",
                          "Stop Time Steam",
                          "Baseline Steam"]


#Some prelin
baseline_by_day_to_print_elec=[]
for i in range(len(baseline_by_day_elec)):
    baseline_by_day_to_print_elec.append(baseline_by_day_elec[i][0])

baseline_by_day_to_print_steam=[]
for i in range(len(baseline_by_day_steam)):
    baseline_by_day_to_print_steam.append(baseline_by_day_steam[i][0])

#Put together data to print
operating_hours_data=[ts_year_of_days,
                      wbt_daily_ave,
                      start_time_each_day_elec,
                      end_time_each_day_elec,
                      baseline_by_day_to_print_elec,
                      start_time_each_day_steam,
                      end_time_each_day_steam,
                      baseline_by_day_to_print_steam]

#Print the data to file
for i in range(len(operating_hours_headings)):
    c2=ws_oper.cell(row=0,column=i)
    c2.value=operating_hours_headings[i]

    for j in range(len(operating_hours_data[i])):
        c2=ws_oper.cell(row=j+1,column=i)
        c2.value=operating_hours_data[i][j]




##-----------------------------------------------------------------------------------
## Printing the single day stat results

#Make Sheet
ws3=wb.create_sheet(-1,"Single Day Stats")

#Headings
single_day_stat_headings=["Average WkDay Elec",
                          "Average WkEnd Elec",
                          "Peak Day Elec",
                          "Average WkDay Steam",
                          "Average WkEnd Steam",
                          "Peak Day Steam"]

#Data
single_day_stat_data=[wk_day_average_for_date_range_elec,
                      wk_end_average_for_date_range_elec,
                      peak_day_for_date_range_elec,
                      wk_day_average_for_date_range_steam,
                      wk_end_average_for_date_range_steam,
                      peak_day_for_date_range_steam]

#Print
for i in range(len(single_day_stat_headings)):
    c3=ws3.cell(row=0,column=i)
    c3.value=single_day_stat_headings[i]

    for j in range(len(single_day_stat_data[i])):
        c3=ws3.cell(row=j+1,column=i)
        c3.value=single_day_stat_data[i][j]

#For this Tab I need to print some extra stuff. 
c3=ws3.cell(row=0, column=(len(single_day_stat_data)))
c3.value="Peak Date Elec"

c3=ws3.cell(row=1, column=(len(single_day_stat_data)))
c3.value=peak_date_for_date_range_elec

c3=ws3.cell(row=0, column=(len(single_day_stat_data)+1))
c3.value="Peak Date Steam"

c3=ws3.cell(row=1, column=(len(single_day_stat_data)+1))
c3.value=peak_date_for_date_range_steam


#----------------------------------------------------------------------------------------------
wb.save(output_book)

##--------------------------------------------------------
time_list.append(time.time())
print str(round(time_list[-1]-time_list[-2],1))+" seconds"
##--------------------------------------------------------

print "Use single day stat results to estimate start and stop times: "

bucket_date_range=wam.get_bucket_date_range_from_user()

start_date_index=date_list.index(bucket_date_range[0])

end_date_index=date_list.index(bucket_date_range[1])

bucket_open_closed_elec=wam.get_operating_hours_from_user()


##--------------------------------------------------------
time_list.append(time.time())
#print str(round(time_list[-1]-time_list[-2],1))+" seconds"
##--------------------------------------------------------

print "Bucketing usage and printing results to new book.",

bucket_operating_hours_by_day_elec=[]
for i in range((bucket_date_range[1]-bucket_date_range[0]).days):
    bucket_operating_hours_by_day_elec.append(bucket_open_closed_elec)


bucket_open_closed_steam=wam.get_operating_hours_from_user()

bucket_operating_hours_by_day_steam=[]
for i in range((bucket_date_range[1]-bucket_date_range[0]).days):
    bucket_operating_hours_by_day_steam.append(bucket_open_closed_steam)

bucketed_usage_elec=wam.get_bucketed_usage(bucket_operating_hours_by_day_elec, date_list, start_date_index, end_date_index,
                       interval_usage_by_day_elec)

bucketed_usage_steam=wam.get_bucketed_usage(bucket_operating_hours_by_day_steam, date_list, start_date_index, end_date_index,
                       interval_usage_by_day_steam)

#printing usage to new excel sheet
output_book_buckets=chan.add_to_filename(book_name," - Bucketed Usage - "+str(int(time_list[0])))

wb_buckets = Workbook()

ws_buckets=wb_buckets.create_sheet(0,"Bucketed Usage Elec")

bucket_headings=["Date", "Usage Open Hours Elec", "Usage Closed Hours Elec", "Usage Open Hours Steam", "Usage Closed Hours Steam"]

bucket_data=[bucketed_usage_elec[2], bucketed_usage_elec[0], bucketed_usage_elec[1], bucketed_usage_steam[0], bucketed_usage_steam[1]]

for i in range(len(bucket_headings)):
    c=ws_buckets.cell(row=0,column=i)
    c.value=bucket_headings[i]

    ## for all rows j+1
    for j in range(len(bucket_data[i])):
        c=ws_buckets.cell(row=j+1,column=i)
        c.value=bucket_data[i][j]

wb_buckets.save(output_book_buckets)

##--------------------------------------------------------
time_list.append(time.time())
print str(round(time_list[-1]-time_list[-2],1))+" seconds"
##--------------------------------------------------------


raw_input("Press any key to exit, I prefer enter, don't forget to update the buckets")

print "Exited Program"


##bucket_closed_usage=[]
##bucket_open_usage=[]
##
##intermediate_week_open=0
##intermediate_week_closed=0
##
##for i in range(start_date_index, end_date_index):
##    
##    if date_list[i].isoweekday()<6:
##
##        for j in range(len(bucket_operating_hours_by_day[i-start_date_index])):
##            if bucket_operating_hours_by_day[i-start_date_index][j]==1:
##                intermediate_week_open+=interval_usage_by_day_elec[i][j]
##            else:
##                intermediate_week_closed+=interval_usage_by_day_elec[i][j]
##
##    if date_list[i].isoweekday()==7:
##        
##        bucket_open_usage.append(intermediate_week_open)
##        bucket_closed_usage.append(intermediate_week_closed)
##        
##        intermediate_week_open=0
##        intermediate_week_closed=0
        
            
    







################ ----------------------------Plotting Suff-------------------------------------
##############print "Starting plot module, exit graph and go to interpreter to plot another day"
##############
##############exit_flag="N"
###############exit_flag=0
###############day_of_year=0
##############date_of_year="12/25/2013"
##############index=50
##############while exit_flag!="Y":
##############    
##############    steam_or_elec=raw_input("elec[0] or steam[1]")
##############
##############    if steam_or_elec=='0':
##############
##############        try:
##############
##############            plot_band=pl.plot_date(interval_time_by_day_elec[index],interval_usage_by_day_elec[index],'g-')
##############            plot_band=pl.plot_date(interval_time_by_day_elec[index],year_of_std_upper_elec[index],'b-')
##############            plot_band=pl.plot_date(interval_time_by_day_elec[index],year_of_std_lower_elec[index],'r-')
##############            plot_band=pl.plot_date(interval_time_by_day_elec[index],baseline_by_day_elec[index],'y')
##############            
##############            xaxisdate=interval_time_by_day_elec[index][0]
##############            
##############            xaxislabel=(
##############                           "Year:"         + str(xaxisdate.year)
##############                        + " Month:"        + str(xaxisdate.month)
##############                        + " Day of month:" + str(xaxisdate.day)
##############                        + " Day of week:"  + str(xaxisdate.isoweekday())
##############                           )
##############            try:
##############                xaxislabel+=" Start Time:" + str(start_time_each_day_elec[index].hour)+":"+str(start_time_each_day_elec[index].minute)
##############            except:
##############                xaxislabel+=" Start Time: Not Found"
##############
##############            try:
##############                xaxislabel+=" End Time:"   + str(end_time_each_day_elec[index].hour)+":"+str(end_time_each_day_elec[index].minute)
##############            except:
##############                xaxislabel+=" End Time: Not Found occured next day"
##############            
##############            
##############            plot_band=pl.xlabel(xaxislabel)
##############            pl.show()
##############            print "Got to end of plot try block"
##############
##############        except:
##############            print "Are you sure you entered either 'Y' or an int from 1 to MAX, inclusive?"
##############
##############        exit_flag=raw_input(["Y to exit or num from 1 to MAX to plot a day"])
##############        #exit_flag=exit_flag+1
##############        try:
##############            index=int(exit_flag)-1
##############        except:
##############            index=exit_flag
##############
##############    elif steam_or_elec=='1':
##############
##############        try:
##############
##############            plot_band=pl.plot_date(interval_time_by_day_steam[index],interval_usage_by_day_steam[index],'g-')
##############            plot_band=pl.plot_date(interval_time_by_day_steam[index],year_of_std_upper_steam[index],'b-')
##############            plot_band=pl.plot_date(interval_time_by_day_steam[index],year_of_std_lower_steam[index],'r-')
##############            plot_band=pl.plot_date(interval_time_by_day_steam[index],baseline_by_day_steam[index],'y')
##############            
##############            xaxisdate=interval_time_by_day_steam[index][0]
##############            
##############            xaxislabel=(
##############                           "Year:"         + str(xaxisdate.year)
##############                        + " Month:"        + str(xaxisdate.month)
##############                        + " Day of month:" + str(xaxisdate.day)
##############                        + " Day of week:"  + str(xaxisdate.isoweekday())
##############                           )
##############            try:
##############                xaxislabel+=" Start Time:" + str(start_time_each_day_steam[index].hour)+":"+str(start_time_each_day_steam[index].minute)
##############            except:
##############                xaxislabel+=" Start Time: Not Found"
##############
##############            try:
##############                xaxislabel+=" End Time:"   + str(end_time_each_day_steam[index].hour)+":"+str(end_time_each_day_steam[index].minute)
##############            except:
##############                xaxislabel+=" End Time: Not Found occured next day"
##############            
##############            
##############            plot_band=pl.xlabel(xaxislabel)
##############            pl.show()
##############            print "Got to end of plot try block"
##############
##############        except:
##############            print "Are you sure you entered either 'Y' or an int from 1 to MAX, inclusive?"
##############
##############        exit_flag=raw_input(["Y to exit or num from 1 to MAX to plot a day"])
##############        #exit_flag=exit_flag+1
##############        try:
##############            index=int(exit_flag)-1
##############        except:
##############            index=exit_flag
##############
##############    else:
##############        print "Try again"
##############


import datetime, os, wam, time
import numpy as np
import pylab as pl

from openpyxl import Workbook
from openpyxl import load_workbook

## The goal is now to print to excel all the results.
## I also want to make the program pull from a single source sheet.
## what do I want to output though?
## This takes weather data and picks the similar days, so on the weather side, I want to output the similar days
## for every hour
## On the energy side, I want to output the average, standard dev, up, down, and similar days used for each interval. 

## Important Declarations
holidays=[2,                                        ## New Day
         16,                                        ## MLK Day
         31+20,                                     ## Pres Day
         31+29+31+30+28,                            ## Mem Day
         31+29+31+30+31+30+4,                       ## Indy Day
         31+29+31+30+31+30+31+31+3,                 ## Lab Day
         31+29+31+30+31+30+31+31+30+8,              ## Col Day
         31+29+31+30+31+30+31+31+30+31+12,          ## Vets Day
         31+29+31+30+31+30+31+31+30+31+22,          ## Thanks Day
         31+29+31+30+31+30+31+31+30+31+23,          ## Coma Day
         31+29+31+30+31+30+31+31+30+31+30+25        ## Christ Day
        ]


## How many similar days do you want to return?
num_matches=6

## Boring prelim stuff
working_directory="C:/Users/James McGlynn/Documents/GitHub/Utility Analysis/Weather/"
book_name="Hourly_Wetbulb.xlsx"
weather_book_path=working_directory+book_name

time_list=[]

time_list.append(time.time())
## ---------------------------Get data from spreadsheet-----------------------------

print "Retrieving weather data from "+book_name+": ",

## This function returns a list of two lists
tstamp_wbulb=wam.xlsx2np(weather_book_path)

## The first is hourly time stamp
time_stamp_np=tstamp_wbulb[0]

## The second is hourly wetbulb temp
wbt_np=tstamp_wbulb[1]

## Use the data from the spreadsheet to define a list that has the date starting from start date to end date
## In case there are gaps. At the moment the data has to be in date order oldest to newest. 
current_date=time_stamp_np[0]
end_date=time_stamp_np[-1]
date_list=[]
while current_date<=end_date:
    date_list.append(current_date)
    current_date=current_date+datetime.timedelta(days=1)

time_list.append(time.time())
print str(round(time_list[-1]-time_list[-2],1))+" seconds"

## ------------------Get the average wet bulb temperature for each day---------------
print "Finding the average wetbulb temperature for each day: ",

## This function also returns a list of two lists
by_day=wam.interval2day(time_stamp_np,wbt_np)

## One list of the time stamp BY DAY 
ts_by_day=by_day[0]

## One list of the wetbulb temp BY DAY
wbt_by_day=by_day[1]

## This function days a list of lists and returns a list of averages
wbt_daily_ave=wam.list_of_lists_2_list_of_ave(wbt_by_day)

time_list.append(time.time())
print str(round(time_list[-1]-time_list[-2],1))+" seconds"

## ---------------------Find the N most similar days to each day--------------------
print "Finding similar days based on a single criteria (Average WBT): ",

## This function takes a list of numbers, the number of closest matches to get, and a date list
## The date list is used as a criteria because I actually only want to consider
## Similar WB temps that occur on the same day of the week. 
similar_days_by_day=wam.get_n_closest_matches_for_each_item_in_list(wbt_daily_ave, num_matches, date_list, holidays)


time_list.append(time.time())
print str(round(time_list[-1]-time_list[-2],1))+" seconds"

## ---------------------Get interval usage data---------------------

print "Retrieving interval data from spreadsheet: ",

interval_book_path="C:/Users/James McGlynn/My Programs/Python Programs/Utility Data Analysis/Interval Data/interval_data_2012.xlsx"

## This function returns a list of two lists
interval_data=wam.xlsx2np(interval_book_path)

## The first is interval time stamp
interval_time=interval_data[0]

## The second is interval data (in this case steam)
interval_usage=interval_data[1]

#---------------------------------------------------------------
time_list.append(time.time())
print str(round(time_list[-1]-time_list[-2],1))+" seconds"
#---------------------------------------------------------------


print "Breaking up interval data by day: ",

## This function also returns a list of two lists
interval_by_day=wam.interval2day(interval_time,interval_usage)

## One list of the time stamp BY DAY 
interval_time_by_day=interval_by_day[0]

## One list of the wetbulb temp BY DAY
interval_usage_by_day=interval_by_day[1]

time_list.append(time.time())
print str(round(time_list[-1]-time_list[-2],1))+" seconds"

##---------------------Apply weather findings to interval data---------------------

print "Gathering interval usage for days that had similar weather: ",

similar_days_interval_usage=wam.use_list_of_list_of_indices_to_group_a_list_of_lists(interval_usage_by_day,similar_days_by_day)

##--------------------------------------------------------
time_list.append(time.time())
print str(round(time_list[-1]-time_list[-2],1))+" seconds"
##--------------------------------------------------------

##----------------------Do simple stat calcs------------------------------------

print "Reorganizing the data and calculating average and stdev for each interval value: ",
## Reorganizing the data to make it easier to get average days and calculate std and stuff
similar_days_interval_usage_by_interval=wam.zip_all_items_of_a_list(similar_days_interval_usage)

## This function returns a list with three lists that are the same shape as orig list except the deepest list turned into a number
stats=wam.get_ave_std_of_list_of_list_of_list(similar_days_interval_usage_by_interval)
## Average
year_of_average_days=stats[0]
## Upper bound
year_of_std_upper=stats[1]
## Lower bound
year_of_std_lower=stats[2]
## STD
year_of_std=stats[3]


##--------------------------------------------------------
time_list.append(time.time())
print str(round(time_list[-1]-time_list[-2],1))+" seconds"
##--------------------------------------------------------

##--------------------------------------------------------
print "The total runtime to this point was: "+str(round(time_list[-1]-time_list[0],1))+" seconds"
##--------------------------------------------------------

##-----------------------------Printing Shit to Excel------------------------------

print "Printing results to excel: ",
## Pick excel destination book name

interval_averages=[]
interval_upper_bound=[]
interval_lower_bound=[]
interval_std=[]
for i in range(len(year_of_average_days)):
    for j in range(len(year_of_average_days[i])):
        interval_averages.append(year_of_average_days[i][j])
        interval_upper_bound.append(year_of_std_upper[i][j])
        interval_lower_bound.append(year_of_std_lower[i][j])
        interval_std.append(year_of_std[i][j])



output_book=working_directory+"interval_analysis_results.xlsx"


column_headings=["Time Stamp","Mlbs Steam","Average Mlbs Steam","Stdev","Ave+Std","Ave-Std"]
output_list=[interval_time,interval_usage,interval_averages,interval_std,interval_upper_bound,interval_lower_bound]

wb = Workbook()


#ws=wb.get_active_sheet()
ws1=wb.create_sheet(0,"Interval Analysis")
## for all headings i
for i in range(len(column_headings)):
    c=ws1.cell(row=0,column=i)
    c.value=column_headings[i]

    ## for all rows j+1
    for j in range(len(output_list[i])):
        c=ws1.cell(row=j+1,column=i)
        c.value=output_list[i][j]



ts_year_of_days=[]
day_of_year=[]
for day_of_hours in ts_by_day:
    try:
        ts_year_of_days.append(day_of_hours[0])
        day_of_year.append(day_of_hours[0].timetuple()[7])
    except:
        ts_year_of_days.append("err")
        day_of_year.append("err")
                    
similar_days_by_day_string_list=[]
for day in similar_days_by_day:
    inter_string=""
    for sim_day in day:
        inter_string=inter_string+str(sim_day)+","
        
    similar_days_by_day_string_list.append(inter_string)


day_anal_headings=["Time Stamp","Day of Year","Wetbulb Temp","Similar Days"]
output_list_by_day=[ts_year_of_days,day_of_year,wbt_daily_ave,similar_days_by_day_string_list]

ws2=wb.create_sheet(-1,"Day Analysis")

for i in range(len(day_anal_headings)):
    c2=ws2.cell(row=0,column=i)
    c2.value=day_anal_headings[i]

    for j in range(len(output_list_by_day[i])):
        c2=ws2.cell(row=j+1,column=i)
        c2.value=output_list_by_day[i][j]
    
    
wb.save(output_book)

##--------------------------------------------------------
time_list.append(time.time())
print str(round(time_list[-1]-time_list[-2],1))+" seconds"
##--------------------------------------------------------

## ----------------------------Plotting Suff-------------------------------------
print "Starting plot module, exit graph and go to interpreter to plot another day"

exit_flag="N"
#exit_flag=0
day_of_year=0
while exit_flag!="Y":
    
    try:
        day_of_year_time=interval_time_by_day[day_of_year]
        day_of_year_usage=interval_usage_by_day[day_of_year]
        day_of_year_stdup=year_of_std_upper[day_of_year]
        day_of_year_stdlo=year_of_std_lower[day_of_year]

        steam_plot_band=pl.plot_date(day_of_year_time,day_of_year_usage,'g-')
        steam_plot_band=pl.plot_date(day_of_year_time,day_of_year_stdup,'b-')
        steam_plot_band=pl.plot_date(day_of_year_time,day_of_year_stdlo,'r-')
        
        xaxisdate=day_of_year_time[0]
        xaxislabel="Year:"+str(xaxisdate.year)+" Month:"+str(xaxisdate.month)+ " Day of month:"+str(xaxisdate.day) + " Day of week:"+str(xaxisdate.isoweekday())
        steam_plot_band=pl.xlabel(xaxislabel)
        pl.show()

    except:
        "Are you sure you entered either 'Y' or an int from 1 to 366, inclusive?"

    exit_flag=raw_input(["Y to exit or num from 1 to 366 to plot a day"])
    #exit_flag=exit_flag+1
    try:
        day_of_year=int(exit_flag)-1
    except:
        day_of_year=exit_flag

print "Exiting Program"







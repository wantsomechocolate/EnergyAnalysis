
from openpyxl import Workbook
from openpyxl import load_workbook
import numpy as np
import datetime, array
from dateutil import parser


def dtobj2xl(date_obj_def):
    base_day_def=datetime.datetime(1899,12,30)
    elapsed_def=date_obj_def-base_day_def
    xltime_def=elapsed_def.total_seconds()/(24*3600)
    return xltime_def
    

def xl2dtobj(xl_date_def,datemode_def):
    return (datetime.datetime(1899,12,30)+ datetime.timedelta(days=xl_date_def + 1462 * datemode_def))


def xlsx2np(book_path_def, sheet_name_def, num_col_def):

    wb=load_workbook(book_path_def)

    ## Load the active sheet (only one sheet)
    ws = wb.get_sheet_by_name(sheet_name_def)

    ## Get the last occupied row of data - the code assumed the number of columns and what is in them
    last_occ_row_def=ws.rows[-1][0].row

    ## Initialize some stuff
    #time_stamp_list_def=[]
    #wbt_list_def=[]

    return_list_def=[]
    for k in range(num_col_def):
        return_list_def.append([])

    ## Get every item in the spreadsheet (thousands at this point
    ## and put the datestamp and wetbulb temp into a numpy array. 
    for i in range(last_occ_row_def):
        for j in range(len(return_list_def)):
            
            c_def=ws.cell(row=i, column=j)
            return_list_def[j].append(c_def.value)         

    #time_stamp_np=np.array(time_stamp_list[1:])
    #wbt_np=np.array(wbt_list[1:])

    ## this removes the heading name - I should really make this a keyed list and use the
    ## column heading as the key!
    for i in range(len(return_list_def)):
        return_list_def[i]=return_list_def[i][1:]
    
    #ts_wbt=[time_stamp_np,wbt_np]

    return return_list_def

def xlsx2np_v2(book_path_def, sheet_name_def, num_col_def):

    wb=load_workbook(book_path_def)

    ## Load the active sheet (only one sheet)
    ws = wb.get_sheet_by_name(sheet_name_def)

    data_list_addresses_def=ws.columns

    verts=array.array('i',(0,)*len(ws.rows))
    data_range_def=[]
    for i in range(len(ws.columns)):
        data_range_def.append(verts)

    #data_list_addresses_def = zip(*data_list_addresses_def)

    for i in range(len(data_list_addresses_def)):
        for j in range(len(data_list_addresses_def[i])):
            data_range_def[i][j]=data_list_addresses_def[i][j].value

    return data_range_def
            

######    ## Get the last occupied row of data - the code assumed the number of columns and what is in them
######    last_occ_row_def=ws.rows[-1][0].row
######
######    ## Initialize some stuff
######    #time_stamp_list_def=[]
######    #wbt_list_def=[]
######
######    return_list_def=[]
######    for k in range(num_col_def):
######        return_list_def.append([])
######
######    ## Get every item in the spreadsheet (thousands at this point
######    ## and put the datestamp and wetbulb temp into a numpy array. 
######    for i in range(last_occ_row_def):
######        for j in range(len(return_list_def)):
######            
######            c_def=ws.cell(row=i, column=j)
######            return_list_def[j].append(c_def.value)         
######
######    #time_stamp_np=np.array(time_stamp_list[1:])
######    #wbt_np=np.array(wbt_list[1:])
######
######    ## this removes the heading name - I should really make this a keyed list and use the
######    ## column heading as the key!
######    for i in range(len(return_list_def)):
######        return_list_def[i]=return_list_def[i][1:]
######    
######    #ts_wbt=[time_stamp_np,wbt_np]

    #return return_list_def


def interval2day(time_list_def, data_list_def):
    ## Set first date to its own variable
    
    #start_date_def=time_list_def[0]
    #end_date_def=time_list_def[-1]
    
    #elap_td_def=end_date_def-start_date_def
    #elap_int_def=elap_td_def.days+1

    ## Initialize a place to hold the wbt by day. 
    #wbt_by_day_def=[]
    #for i in range(elap_int_def):
    #    wbt_by_day_def.append([])

    #ts_by_day_def=[]
    #for i in range(elap_int_def):
    #    ts_by_day_def.append([])


    current_date_def=time_list_def[0]
    end_date_def=time_list_def[-1]
    date_list_def=[]
    while current_date_def<=end_date_def:
        date_list_def.append(current_date_def)
        current_date_def=current_date_def+datetime.timedelta(days=1)

    ts_by_day_def=[]
    for i in range(len(date_list_def)):
        ts_by_day_def.append([])
    
    wbt_by_day_def=[]
    for i in range(len(date_list_def)):
        wbt_by_day_def.append([])


    ## Go through the huge list and put everything where it goes.

    for i in range(len(time_list_def)): ## FOR EVERY SINGLE DATA POINT (8000 FOT TEMP, 30000 FOR INTERVAL)

        for j in range(len(date_list_def)): ## FOR EVERY DAY THAT FALLS IN THE RANGE BETWEEN THE OLDEST AND NEWEST DATE IN THE DATE RANGE

            ## STRIP OFF THE TIME COMPONENT FROM THE INTERVAL DATETIME, IF THE DATE MATCHES THE CURRENT DAY THEN ADD, IF NOT MOVE ON!
            interval_data_day_def=datetime.datetime(time_list_def[i].year,time_list_def[i].month,time_list_def[i].day)

            if interval_data_day_def==date_list_def[j]:
                   ts_by_day_def[j].append(time_list_def[i])
                   wbt_by_day_def[j].append(data_list_def[i])

                   
        
        #for j in range(elap_int_def):

        #    ## this if statement needs to stop doing this right now! to be fully robust it should do a year-month-day comparison
        #    if time_list_def[i].timetuple().tm_yday-1==j: ######WWWWWRRRRRROOOOONNNNNGGGGGGGGGGGGGGG
        #        wbt_by_day_def[j].append(data_list_def[i])
        #        ts_by_day_def[j].append(time_list_def[i])

    return_list_def=[ts_by_day_def,wbt_by_day_def]

    return return_list_def

                   

def list_of_lists_2_list_of_ave(list_of_lists_def):

    list_of_aves_def=[]
    for i in range(len(list_of_lists_def)):
        try:
            daily_ave_def=float(sum(list_of_lists_def[i]))/len(list_of_lists_def[i])
        except:
            daily_ave_def="err"
        list_of_aves_def.append(daily_ave_def)
    return list_of_aves_def


#This function is working just fine, 6 closest matches is too many for one year of data, it looks like 2 or three would be better.
# I'm going to use 4 for now. I still want to find out why the averages don't exactly match and I still need to fix the date issue.
def get_n_closest_matches_for_each_item_in_list(list_of_nums_def,n_count_def,criteria_date_def,exclude_days_def):
    ## What is N?
    #min_count=n_count
    
    indices_of_matches_def=[]

    ## For each item in the list that this thing is supposed to find the n closest matches for. 
    for i in range(len(list_of_nums_def)):

        day_of_year_def=[]
        difference_def=[] 
        min_indices_def=[]
        diff_list_def=[]


        ## For each of the closest days that this function is supposed to find
        for n in range(n_count_def):

            ## This is essentially making a list of 0's with the lenth to fit all the closest days
            min_indices_def.append(0)

        ## Iterate again through the list so that each item can be compared with every other item
        for j in range(len(list_of_nums_def)):

            ## If the current config is good
            if (criteria_date_def[i].isoweekday()==criteria_date_def[j].isoweekday()) and (i!=j) and (criteria_date_def[j] not in exclude_days_def):

                ## append the list index of the wetbulb that is a potential candidate
                day_of_year_def.append(j)

                ## try to get the absolute value of the difference between metrics for comparison
                try:
                    difference_def.append(abs(list_of_nums_def[i]-list_of_nums_def[j]))
                except:
                    difference_def.append("err")

        ## populate diff_list with what is described above. diff list will be a list of two lists
        ## of the same lenth - one with indices in the original master list and one with abs diff
        ## the list for each iteration through days of the year will only contain days that land
        ## on the same day of the week
        diff_list_def=[day_of_year_def,difference_def]
        
        ## For each of the N values I'm supposed to get from the diff list.
        for k in range(len(min_indices_def)):

            ## Find the min (closest value to current days) the first time through, this will likely
            ## BE the current day. except for the fact that I excluded that from being a candidate day above.
            
            min_val_def=min(diff_list_def[1])

            ## then get the index of that minimum value in diff_list[i]
            index_of_min_val_in_diff_list_def=diff_list_def[1].index(min_val_def)

            ## Then get master list index associated with that diff list index
            day_of_year_min_val_occurred_def=diff_list_def[0][index_of_min_val_in_diff_list_def]

            ## then save the INDEX of the min value
            min_indices_def[k]=day_of_year_min_val_occurred_def

            ## Then alter diff_list so that the value at the saved index is no longer even
            ## close to being a match.
            diff_list_def[1][index_of_min_val_in_diff_list_def]="already matched"

        ## Add the indices (a list) to a bigger list that will hold a list of the N closest values
        ## for each day. 
        #min_indices_list.append(min_indices)

        indices_of_matches_def.append(min_indices_def)

    return indices_of_matches_def


## This function takes a list of lists and another list of lists and groups the first list based
## on the indices in the second one?
def use_list_of_list_of_indices_to_group_a_list_of_lists(main_list_def,list_of_list_indices_def):

    main_list_with_criteria_def=[]

    ## for item in sim days by day (each item will be a list of ints corresponding to days aka list indices)
    for i in range(len(list_of_list_indices_def)):

        ## an interim variable to hold a single day?
        interim_list_def=[]

        ## for each list index corresponding to day of year
        for j in range(len(list_of_list_indices_def[i])):

            ## Go to list with the data you want and collect it
            interim_list_def.append(main_list_def[list_of_list_indices_def[i][j]])

        ## put all the days usages collected into the final list
        main_list_with_criteria_def.append(interim_list_def)

        ## the result of this is a list of 366 days, each day has N similar days in it, each similar day has 96 values. 

    return main_list_with_criteria_def

def zip_all_items_of_a_list(list_to_zip_def):
    zipped_list_def=[]
    for item in list_to_zip_def:
        zipped_list_def.append(zip(*item))

    return zipped_list_def

def get_ave_std_of_list_of_list_of_list(list_to_analyze_def):

    list_average_def=[]
    list_std_def=[]
    list_average_plus_stdev_def=[]
    list_average_minus_stdev_def=[]

    for list_of_list_def in list_to_analyze_def:
        inter_list_ave_def=[]
        inter_list_std_def=[]
        inter_list_stdup_def=[]
        inter_list_stdlo_def=[]
        for list_ in list_of_list_def:

            inter_def=np.array(list_)

            
            inter_def_new=[]
            for item in inter_def:
                try:
                    inter_def_new.append(float(item))
                except:
                    pass
            inter_def=np.array(inter_def_new)

            
            inter_list_ave_def.append(inter_def.mean())
            inter_list_std_def.append(inter_def.std())
            inter_list_stdup_def.append(inter_def.mean()+inter_def.std())
            inter_list_stdlo_def.append(inter_def.mean()-inter_def.std())


                

                
        list_average_def.append(inter_list_ave_def)
        list_std_def.append(inter_list_std_def)
        list_average_plus_stdev_def.append(inter_list_stdup_def)
        list_average_minus_stdev_def.append(inter_list_stdlo_def)

    return_list_def=[list_average_def,list_average_plus_stdev_def,list_average_minus_stdev_def,list_std_def]

    return return_list_def

def ceil_to_one_sig_fig(number_to_ceil_def):

    magnitude_def=10**(int(len(str(int(number_to_ceil_def))))-1)
    if number_to_ceil_def%magnitude_def==0:
        
        return number_to_ceil_def
    
    elif number_to_ceil_def<0:
        opp_def=number_to_ceil_def*-1
        temp_ceil_def=floor_to_one_sig_fig_def(opp_def)
        ceil_def=temp_ceil_def*-1
        return ceil_def

    else:

        actual_ceil_def=(int(str(int(number_to_ceil_def))[0:1])+1)*10**(int(len(str(int(number_to_ceil_def))))-1)

        return actual_ceil_def

def floor_to_one_sig_fig(number_to_floor_def):

    if number_to_floor_def<0:
        opp_def=number_to_floor_def*-1
        temp_floor_def=ceil_to_one_sig_fig_def(opp_def)
        floor_def=temp_floor_def*-1
        return floor_def

    else:

        actual_floor_def=(int(str(int(number_to_floor_def))[0:1]))*10**(int(len(str(int(number_to_floor_def))))-1)

        return actual_floor_def

def get_ave_of_k_min_values(list_to_take_mins_from_all_def,num_of_min_values_def,index_start_def, index_end_def):

    ## This is done in case you want to look at a beginning of the day baseline and not have crazy
    ## end of the day stuff get in the way.
    
    list_to_take_mins_from_def=list_to_take_mins_from_all_def[index_start_def:index_end_def]

    if len(list_to_take_mins_from_def)<num_of_min_values_def:
        return np.average(list_to_take_mins_from_def)

    else:
        k_min_values_def=[]
        list_to_take_mins_from_copy_def=list(list_to_take_mins_from_def)
        for i in range(num_of_min_values_def):
            min_value_def=min(list_to_take_mins_from_copy_def)
            list_to_take_mins_from_copy_def.remove(min_value_def)
            k_min_values_def.append(min_value_def)

        try:
            k_min_average_def=np.average(k_min_values_def)
        except:
            k_min_average_def="Not enough min values"

        return k_min_average_def


def get_baseline_by_day(list_of_usages_by_day_def,num_values_def, start_index_def, end_index_def):
    
    baseline_by_day_def=[]
    for i in range(len(list_of_usages_by_day_def)):
        baseline_by_day_def.append([])

    for i in range(len(list_of_usages_by_day_def)):
        baseline_def=get_ave_of_k_min_values(list_of_usages_by_day_def[i],num_values_def, start_index_def, end_index_def)
        for j in range(len(list_of_usages_by_day_def[i])):
            baseline_by_day_def[i].append(baseline_def)

    return baseline_by_day_def



def getholidays():#to_, from_):
    hollydays_def=[
                datetime.datetime(2010,12,31),      ## New Day
                datetime.datetime(2011,1,1),        ## New Day
                datetime.datetime(2011,1,17),       ## MLK Day
                datetime.datetime(2011,2,21),       ## Pres Day
                datetime.datetime(2011,5,30),       ## Mem Day
                datetime.datetime(2011,7,4),        ## Indy Day
                datetime.datetime(2011,9,5),        ## Lab Day
                datetime.datetime(2011,10,10),       ## Col Day
                datetime.datetime(2011,11,11),      ## Vets Day
                datetime.datetime(2011,11,24),      ## Thanks Day
                datetime.datetime(2011,11,25),      ## Coma Day
                datetime.datetime(2011,12,26),      ## Christ Day
        
                datetime.datetime(2012,1,2),        ## New Day
                datetime.datetime(2012,1,16),       ## MLK Day
                datetime.datetime(2012,2,20),       ## Pres Day
                datetime.datetime(2012,5,28),       ## Mem Day
                datetime.datetime(2012,7,4),        ## Indy Day
                datetime.datetime(2012,9,3),        ## Lab Day
                datetime.datetime(2012,10,8),       ## Col Day
                datetime.datetime(2012,11,12),      ## Vets Day
                datetime.datetime(2012,11,22),      ## Thanks Day
                datetime.datetime(2012,11,23),      ## Coma Day
                datetime.datetime(2012,12,25),      ## Christ Day

                datetime.datetime(2013,1,1),        ## New Day
                datetime.datetime(2013,1,21),       ## MLK Day
                datetime.datetime(2013,2,18),       ## Pres Day
                datetime.datetime(2013,5,27),       ## Mem Day
                datetime.datetime(2013,7,4),        ## Indy Day
                datetime.datetime(2013,9,2),        ## Lab Day
                datetime.datetime(2013,10,14),      ## Col Day
                datetime.datetime(2013,11,11),      ## Vets Day
                datetime.datetime(2013,11,28),      ## Thanks Day
                datetime.datetime(2013,11,29),      ## Coma Day
                datetime.datetime(2013,12,25)       ## Christ Day
                ]
    
    return hollydays_def

def get_start_time_each_day(interval_time_by_day_def,interval_usage_by_day_def, baseline_by_day_def, percent_above_baseline_def, threshold_def):

    ## for every day
    start_time_each_day_def=[]
    
    for i in range(len(interval_usage_by_day_def)):

        found_start_time_def="No"
        start_time_def="default value"
        count=0

        ## for every 15 minute period

        if "Not enough min values" in baseline_by_day_def[i]:
            start_time_each_day_def.append("Not enough min values")
        else:
        
            for j in range(len(interval_usage_by_day_def[i])):
                

                ##if the usage during the 15 minutes is greater than a percentage more than the baseline
                if interval_usage_by_day_def[i][j]>baseline_by_day_def[i][j]*(1+percent_above_baseline_def):
                    ## increase count
                    count=count+1
                else:
                    ## otherwise reset count
                    count=0

                ## at this point, count could be from 0 to 96 (every value higher than baseline)

                ## if at any point it becomes greater than thresh and it hasn't before today
                if count>=threshold_def and found_start_time_def=="No":
                    ## say that the start time occured when the increase started
                    start_time_def=interval_time_by_day_def[i][j-threshold_def]

                    ## and say that a start time was found for the day
                    found_start_time_def="Yes"

            if found_start_time_def=="Yes":
                if start_time_def.hour==0:
                    start_time_each_day_def.append("err")
                else:
                    start_time_each_day_def.append(start_time_def)
            else:
                start_time_each_day_def.append("N/A")


    return start_time_each_day_def


def get_end_time_each_day(interval_time_by_day_def, interval_usage_by_day_def, baseline_by_day_def, start_time_each_day_def, percent_above_baseline_def, thresh_end_def):

    end_time_each_day_def=[]
    
    start_time_each_day_copy=list(start_time_each_day_def)

    ## for every day
    for i in range(len(interval_usage_by_day_def)):

        ## Refresh some vars
        found_end_time_def="No"
        end_time_def="default val"
        count=0

        if "Not enough min values" in baseline_by_day_def[i]:
            end_time_each_day_def.append("Not enough min values")
        else:


            ## for every 15 minute period
            for j in range(len(interval_usage_by_day_def[i])):


                try:
                    test=start_time_each_day_copy[i].hour
                except:
                    start_time_each_day_copy[i]=interval_time_by_day_def[i][40]

                ## If the 15 minute period in question is not even passed the start time
                if interval_time_by_day_def[i][j]<=start_time_each_day_copy[i]:

                    ## Then don't do anything. 
                    pass

                ## Otherwise, begin/continue the analysis. 
                else:
                
                    ##if the usage during the 15 minutes is less than a percentage more than the baseline
                    if interval_usage_by_day_def[i][j]<baseline_by_day_def[i][j]*(1+percent_above_baseline_def):
                        ## increase count
                        count=count+1
                    else:
                        ## otherwise reset count
                        count=0

                    ## at this point, count could be from 0 to 96 minus the number of points that fell before the start time

                    ## if at any point it becomes greater than or equal to thresh and it hasn't before today
                    if count>=thresh_end_def and found_end_time_def=="No":
                        
                        ## say that the end time occured when thresh was met
                        end_time_def=interval_time_by_day_def[i][j-thresh_end_def]

                        ## and say that an end time was found for the day
                        found_end_time_def="Yes"

            if found_end_time_def=="Yes":
        ##            if end_time.hour==0:
        ##                end_time_each_day.append("err")
        ##            else:
                end_time_each_day_def.append(end_time_def)
            else:
                ## the case that the program gets here on the last element of the array being indexed by i
                ## needs to be coded for. 
                ## If it fails it most likely means that the end time occurs the next day - at least for the data set that I'm using.

                try:

                    for m in range(len(interval_usage_by_day_def[i+1])): #next day
                        if found_end_time_def=="No":
                            if interval_usage_by_day_def[i+1][m]<baseline_by_day_def[i+1][m]*(1+percent_above_baseline_def):
                                found_end_time_def="Yes"
                                end_time_def=interval_time_by_day_def[i+1][m]

                ## Last Day no Shutdown
                except:
                    end_time_def="LDNSD"
                    
                end_time_each_day_def.append(end_time_def)

    return end_time_each_day_def


def get_date_range_from_user(debug_mode):

    if debug_mode==False:
        from dateutil import parser

        got_to_end=False
        while got_to_end==False:
            try:
                start_date=raw_input("Start Date - All common formats are fine, use full year: ")
                sd_obj=parser.parse(start_date)
                got_to_end=True
            except:
                print "Date format not recognized or date does not exist"

        got_to_end=False
        while got_to_end==False:
            try:
                end_date=raw_input("End Date - All common formats are fine, use full year: ")
                ed_obj=parser.parse(end_date)
                if sd_obj<ed_obj:
                    got_to_end=True
                else:
                    print "What! enter the dates in order, fool!"
            except:
                print "Date format not recognized or date does not exist"

        return [sd_obj, ed_obj]

    else:
        return [datetime.datetime(2013,4,1), datetime.datetime(2013,7,1)]

def get_stats_by_day_in_range(interval_usage_by_day_def, date_list_def, date_range_def):
    
    start_index_def=date_list_def.index(date_range_def[0])
    end_index_def=date_list_def.index(date_range_def[1])

    #interval_by_day_in_range_def=interval_by_day_def[start_index_def:end_index_def]


    date_list_wkday_def=[]
    date_list_wkend_def=[]
    interval_wkday_def=[]
    interval_wkend_def=[]

    max_value_reached_def=0

    #print start_index_def
    #print end_index_def
    
    for i in range(start_index_def, end_index_def):
        
        if date_list_def[i].weekday()<=4:
            date_list_wkday_def.append(date_list_def[i])
            interval_wkday_def.append(interval_usage_by_day_def[i])
        else:
            date_list_wkend_def.append(date_list_def[i])
            interval_wkend_def.append(interval_usage_by_day_def[i])


        max_val_by_day_def=max(interval_usage_by_day_def[i])
        if max_val_by_day_def>max_value_reached_def:
            max_value_reached_def=max_val_by_day_def
            peak_day_usage_def=interval_usage_by_day_def[i]
            peak_date_def=date_list_def[i]
        else:
            pass
        


    interval_wkday_zipped_def=zip(*interval_wkday_def)
    interval_wkend_zipped_def=zip(*interval_wkend_def)


    wkday_ave_def=[]
    for item in interval_wkday_zipped_def:
        item_np_def=np.array(item)
        item_ave_def=item_np_def.mean()
        wkday_ave_def.append(item_ave_def)

    wkend_ave_def=[]
    for item in interval_wkend_zipped_def:
        item_np_def=np.array(item)
        item_ave_def=item_np_def.mean()
        wkend_ave_def.append(item_ave_def)

    return [wkday_ave_def,wkend_ave_def,peak_day_usage_def,peak_date_def]


def get_bucket_date_range_from_user():
    #bucket_end_date_text="6/30/2013"

    bucket_end_date_text=raw_input("What is the end date of the year you want to use for bucket analysis? >>> ")
    
    bucket_end_date=parser.parse(bucket_end_date_text)

    while bucket_end_date.isoweekday()!=1:
        bucket_end_date=bucket_end_date-datetime.timedelta(days=1)

    bucket_start_date=bucket_end_date-datetime.timedelta(days=364)

    return [bucket_start_date, bucket_end_date]


def get_operating_hours_from_user():

    #bucket_specifying_open_or_closed="open"
    
    bucket_specifying_open_or_closed=raw_input("specifying open range or closed range? >>> ")

    bucket_start_day_text="0:00"
    #bucket_start_time_text="6:00"
    #bucket_end_time_text="18:00"
    bucket_end_day_text="23:45"

    bucket_start_time_text=raw_input("start time hh:mm >>> ")
    bucket_end_time_text=raw_input("end time hh:mm >>> ")

    bucket_start_day=parser.parse(bucket_start_day_text)
    bucket_start_time=parser.parse(bucket_start_time_text)
    bucket_end_time=parser.parse(bucket_end_time_text)
    bucket_end_day=parser.parse(bucket_end_day_text)

    
    bucket_current_time=bucket_start_day
    bucket_open_closed=[]

    while bucket_current_time<=bucket_end_day:
        
        if bucket_current_time<bucket_start_time or bucket_current_time>=bucket_end_time:
            if bucket_specifying_open_or_closed=="open":
                bucket_open_closed.append(0)
            else:
                bucket_open_closed.append(1)
        else:
            if bucket_specifying_open_or_closed=="open":
                bucket_open_closed.append(1)
            else:
                bucket_open_closed.append(0)
                
        bucket_current_time=bucket_current_time+datetime.timedelta(minutes=15)    

    return bucket_open_closed
    



def get_bucketed_usage(bucket_operating_hours_by_day_def, date_list_def, start_date_index_def, end_date_index_def,
                       interval_usage_by_day_def):

    bucket_open_usage_def=[]
    bucket_closed_usage_def=[]
    bucket_date_def=[]
    
    intermediate_week_open=0
    intermediate_week_closed=0

    for i in range(start_date_index_def, end_date_index_def):
        
        if date_list_def[i].isoweekday()<6:

            for j in range(len(bucket_operating_hours_by_day_def[i-start_date_index_def])):
                if bucket_operating_hours_by_day_def[i-start_date_index_def][j]==1:
                    intermediate_week_open+=interval_usage_by_day_def[i][j]
                else:
                    intermediate_week_closed+=interval_usage_by_day_def[i][j]

        if date_list_def[i].isoweekday()==7:

            bucket_date_def.append(date_list_def[i])
            
            bucket_open_usage_def.append(intermediate_week_open)
            bucket_closed_usage_def.append(intermediate_week_closed)
            
            intermediate_week_open=0
            intermediate_week_closed=0

    return [bucket_open_usage_def, bucket_closed_usage_def, bucket_date_def]



















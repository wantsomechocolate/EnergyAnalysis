import datetime
import os, time #wam
import numpy as np
from marbles import glass as chan
from openpyxl import Workbook
from openpyxl import load_workbook
import wam4 as wam
import wamo4 as wamo
from dateutil import parser
#import pylab as pl

class MyWorkbook(object):

    ## As soon as this class is called...
    def __init__(self, book_name):

        ## Load the workbook
        self.wb = load_workbook(book_name)

        ## Get all the sheet names - the order of this list will be the
        ## same order as the sheets in the book I believe (hope)
        self.sheet_names=self.wb.get_sheet_names()

        self.sheet_objects=[]

        #self.sheet_data_objects=[]
        
        ## For each sheet, go in and get the addresses and actual data for each row and
        ## and column that contain any data. 
        for i in range(len(self.sheet_names)):

            ## This initializes a sheet object, which takes the sheet name
            ## and returns the address of the smallest rectangle that includes
            ## all the data.
            self.sheet_objects.append(MySheet(self.wb, self.sheet_names[i]))

            ## This initializes a sheet data object which takes the addresses and
            ## returns the data.
            #self.sheet_data_objects.append(MySheetData(self.sheet_objects[i].sheet_data_range))

        ## This creates as many objects as sheets with the actual data inside (including the headings)
        self.work_book_data={}
        self.work_book_data_no_headings={}
        for i in range(len(self.sheet_names)):
            self.work_book_data[self.sheet_names[i]]=self.sheet_objects[i].sheet_data

            no_heading_intermediate=[]
            for k in range(len(self.sheet_objects[i].sheet_data)):
                no_heading_intermediate.append(self.sheet_objects[i].sheet_data[k][1:])

            self.work_book_data_no_headings[self.sheet_names[i]]=no_heading_intermediate
            
    
    def get_num_sheets(self):

        return len(self.sheet_objects)

## A sheet object has all the data form the sheet in it.
## Sheet data range is a bunch of cell addresses, sheet data is the data at those addresses.
class MySheet(object):

    def __init__(self, workbook, sheet_name):

        self.sheet=workbook.get_sheet_by_name(sheet_name)

        self.sheet_data_range=self.sheet.columns

        self.sheet_data=[]
        for i in range(len(self.sheet_data_range)):
            self.sheet_data.append([])

        for i in range(len(self.sheet_data_range)):
            for j in range(len(self.sheet_data_range[i])):
                self.sheet_data[i].append(self.sheet_data_range[i][j].value)

class IntervalData(object):

    def __init__(self, interval_data_def, interval_data_by_day_def):

        self.datetime_list=interval_data_def[0]
        
        self.data_list=interval_data_def[1]

        self.datetime_list_by_day=interval_data_by_day_def[0]

        self.data_list_by_day=interval_data_by_day_def[1]

        self.averages_by_day=self.list_of_lists_2_list_of_ave(self.data_list_by_day)

        self.num_matches=5

        self.date_list=self.get_date_list(self.datetime_list)

    def get_elapsed_days(self):
        
        return self.datetime_list_by_day[-1][0]-self.datetime_list_by_day[0][0]


    def list_of_lists_2_list_of_ave(self, list_of_lists_def):

        list_of_aves_def=[]
        
        for i in range(len(list_of_lists_def)):
            try:
                daily_ave_def=float(sum(list_of_lists_def[i]))/len(list_of_lists_def[i])
            except:
                daily_ave_def="err"
            list_of_aves_def.append(daily_ave_def)
        return list_of_aves_def

    def get_date_list(self, datetime_list):

        current_date_def=datetime_list[0]
        end_date_def=datetime_list[-1]
        date_list_def=[]
        while current_date_def<=end_date_def:
            date_list_def.append(current_date_def)
            current_date_def=current_date_def+datetime.timedelta(days=1)

        return date_list_def

        

##    def get_average_weekday(self):
##    def get_average_weekend(self):
##    def get_peak_day(self):
##    def get_start_and_end_time_from_ave_wkday(self):
##    def get_bucketed_usage(self):

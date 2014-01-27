import datetime
import os, time #wam
import numpy as np
from marbles import glass as chan
from openpyxl import Workbook
from openpyxl import load_workbook
import wam4 as wam
import wamo4 as wamo
from dateutil import parser
#import pylab as pl

class MyWorkbook(object):

    ## As soon as this class is called...
    def __init__(self, book_name):

        ## Load the workbook
        print "Loading the workbook"
        self.wb = load_workbook(book_name)

        ## Get all the sheet names - the order of this list will be the
        ## same order as the sheets in the book I believe (hope)
        print "Getting all the sheet names"
        self.sheet_names=self.wb.get_sheet_names()

        self.sheet_objects=[]

        #self.sheet_data_objects=[]
        
        ## For each sheet, go in and get the addresses and actual data for each row and
        ## and column that contain any data.

        print "Getting the cell addresses of all the data in each sheet"
        for i in range(len(self.sheet_names)):

            ## This initializes a sheet object, which takes the sheet name
            ## and returns the addresses of the smallest rectangle that includes
            ## all the data. and also the actual data in a different list
            print "Initializing a sheet object"
            self.sheet_objects.append(MySheet(self.wb, self.sheet_names[i]))

        print "let's see if it even gets to this point."
        
        ## This creates as many objects as sheets with the actual data inside (including the headings)
        self.work_book_data={}
        self.work_book_data_no_headings={}
        for i in range(len(self.sheet_names)):
            self.work_book_data[self.sheet_names[i]]=self.sheet_objects[i].sheet_data

            no_heading_intermediate=[]
            for k in range(len(self.sheet_objects[i].sheet_data)):
                no_heading_intermediate.append(self.sheet_objects[i].sheet_data[k][1:])

            self.work_book_data_no_headings[self.sheet_names[i]]=no_heading_intermediate
            
    
    def get_num_sheets(self):

        return len(self.sheet_objects)

## A sheet object has all the data form the sheet in it.
## Sheet data range is a bunch of cell addresses, sheet data is the data at those addresses.
class MySheet(object):

    def __init__(self, workbook, sheet_name):

        print "Initializing sheet "+str(sheet_name)

        print "getting sheet by name"
        self.sheet=workbook.get_sheet_by_name(sheet_name)

######        print "getting the data range" #This is where it hangs?!
######        self.sheet_data_range=self.sheet.columns


        ##---------------------------------------------------------------------
        ## The above method was hanging here is the new method
        ## Get the "highest row"

        ## I will hard code three columns just to see if this works

        if sheet_name=="Interval Temp":
            row1address,row2address = [],[]
            row1data,row2data = [],[]

            highest_row=self.sheet.get_highest_row()

            print "iterating through cells to get addresses with data"
            
            for i in range(highest_row):
                row1address.append(self.sheet.cell(row=i, column=0))
                row2address.append(self.sheet.cell(row=i, column=1))

            print "iterating through cells with data to get data"
            
            for i in range(len(row1address)):
                row1data.append(row1address[i].value)
                row2data.append(row2address[i].value)

            print "Put it all together"
            self.sheet_data_range=[row1address, row2address]
            self.sheet_data=[row1data, row2data]

        else:
            
            row1address,row2address,row3address = [],[],[]
            row1data,row2data,row3data = [],[],[]
            
            highest_row=self.sheet.get_highest_row()

            print "iterating through cells to get addresses with data"
            
            for i in range(highest_row):
                row1address.append(self.sheet.cell(row=i, column=0))
                row2address.append(self.sheet.cell(row=i, column=1))
                row3address.append(self.sheet.cell(row=i, column=2))

            print "iterating through cells with data to get data"
            
            for i in range(len(row1address)):
                if i%1000==0:
                    print str((i))+" of "+str(highest_row)+": "+str((i*100)/highest_row)+"%"
                row1data.append(row1address[i].value)
                row2data.append(row2address[i].value)
                row3data.append(row3address[i].value)

            print "Put it all together"
            self.sheet_data_range=[row1address, row2address, row3address]
            self.sheet_data=[row1data, row2data, row3data]

            #######_--------------------------------------

    ##        print "initializing an empty list of empty lists for the actual data to live"
    ##        self.sheet_data=[]
    ##        for i in range(len(self.sheet_data_range)):
    ##            self.sheet_data.append([])
    ##
    ##        print "Iterating through the empty list and filling it with the data at each address"
    ##        print "for column of data"
    ##        for i in range(len(self.sheet_data_range)):
    ##            print "for row in that column"
    ##            for j in range(len(self.sheet_data_range[i])):
    ##                print "store the value at that column and row"
    ##                self.sheet_data[i].append(self.sheet_data_range[i][j].value)
                

class IntervalData(object):

    def __init__(self, interval_data_def, interval_data_by_day_def):

        self.datetime_list=interval_data_def[0]
        
        self.data_list=interval_data_def[1]

        self.datetime_list_by_day=interval_data_by_day_def[0]

        self.data_list_by_day=interval_data_by_day_def[1]

        self.averages_by_day=self.list_of_lists_2_list_of_ave(self.data_list_by_day)

        self.num_matches=5

        self.date_list=self.get_date_list(self.datetime_list)

    def get_elapsed_days(self):
        
        return self.datetime_list_by_day[-1][0]-self.datetime_list_by_day[0][0]


    def list_of_lists_2_list_of_ave(self, list_of_lists_def):

        list_of_aves_def=[]
        
        for i in range(len(list_of_lists_def)):
            try:
                daily_ave_def=float(sum(list_of_lists_def[i]))/len(list_of_lists_def[i])
            except:
                daily_ave_def="err"
            list_of_aves_def.append(daily_ave_def)
        return list_of_aves_def

    def get_date_list(self, datetime_list):

        current_date_def=datetime_list[0]
        end_date_def=datetime_list[-1]
        date_list_def=[]
        while current_date_def<=end_date_def:
            date_list_def.append(current_date_def)
            current_date_def=current_date_def+datetime.timedelta(days=1)

        return date_list_def

        

##    def get_average_weekday(self):
##    def get_average_weekend(self):
##    def get_peak_day(self):
##    def get_start_and_end_time_from_ave_wkday(self):
##    def get_bucketed_usage(self):

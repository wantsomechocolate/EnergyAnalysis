
from openpyxl import Workbook
from openpyxl import load_workbook
import numpy as np
import datetime


def xl2dtobj(xl_date,datemode):
    return (datetime.datetime(1899,12,30)+ datetime.timedelta(days=xl_date + 1462 * datemode))


def xlsx2np(book_path):

    wb=load_workbook(book_path)

    ## Load the active sheet (only one sheet)
    ws = wb.get_active_sheet()

    ## Get the last occupied row of data - the code assumed the number of columns and what is in them
    last_occ_row=ws.rows[-1][0].row

    ## Initialize some stuff
    time_stamp_list=[]
    wbt_list=[]

    ## Get every item in the spreadsheet (thousands at this point
    ## and put the datestamp and wetbulb temp into a numpy array. 
    for i in range(last_occ_row):
        c0=ws.cell(row=i, column=0)
        time_stamp_list.append(c0.value)
        
        c1=ws.cell(row=i, column=1)
        wbt_list.append(c1.value)

    time_stamp_np=np.array(time_stamp_list[1:])
    wbt_np=np.array(wbt_list[1:])

    ts_wbt=[time_stamp_np,wbt_np]

    return ts_wbt


def interval2day(time_list, data_list):
    ## Set first date to its own variable 
    start_date=time_list[0]
    end_date=time_list[-1]
    elap_td=end_date-start_date
    elap_int=elap_td.days+1

    ## Initialize a place to hold the wbt by day. 
    wbt_by_day=[]
    for i in range(elap_int):
        wbt_by_day.append([])

    ts_by_day=[]
    for i in range(elap_int):
        ts_by_day.append([])

    ## Go through the huge big temp list and put everything where it goes.
    for i in range(len(time_list)):
        for j in range(elap_int):
            if time_list[i].timetuple().tm_yday-1==j:
                wbt_by_day[j].append(data_list[i])
                ts_by_day[j].append(time_list[i])

    return_list=[ts_by_day,wbt_by_day]

    return return_list

def list_of_lists_2_list_of_ave(list_of_lists):

    list_of_aves=[]
    for i in range(len(list_of_lists)):
        try:
            daily_ave=float(sum(list_of_lists[i]))/len(list_of_lists[i])
        except:
            daily_ave="err"
        list_of_aves.append(daily_ave)
    return list_of_aves



def get_n_closest_matches_for_each_item_in_list(list_of_nums,n_count,criteria_date,exclude_days):
    ## What is N?
    #min_count=n_count
    
    indices_of_matches=[]

    ## For each day
    for i in range(len(list_of_nums)):

        day_of_year=[]
        difference=[] 
        min_indices=[]
        diff_list=[]
        
        for n in range(n_count):
            min_indices.append(0)

        for j in range(len(list_of_nums)):

            if (criteria_date[i].isoweekday()==criteria_date[j].isoweekday()) and (j not in exclude_days):

                day_of_year.append(j)

                try:
                    difference.append(abs(list_of_nums[i]-list_of_nums[j]))
                except:
                    difference.append("err")

        ## populate diff_list with what is described above. diff list will be a list of two lists
        ## of the same lenth - one with indices in the original master list and one with abs diff
        ## the list for each iteration through days of the year will only contain days that land
        ## on the same day of the week
        diff_list=[day_of_year,difference]
        
        ## For each of the N values I'm supposed to get from the diff list.
        for k in range(len(min_indices)):

            ## Find the min (closest value to current days) the first time through, this will likely
            ## BE the current day.
            
            min_val=min(diff_list[1])

            ## then get the index of that minimum value in diff_list[i]
            index_of_min_val_in_diff_list=diff_list[1].index(min_val)

            ## Then get master list index associated with that diff list index
            day_of_year_min_val_occurred=diff_list[0][index_of_min_val_in_diff_list]

            ## then save the INDEX of the min value
            min_indices[k]=day_of_year_min_val_occurred

            ## Then alter diff_list so that the value at the saved index is no longer even
            ## close to being a match.
            diff_list[1][index_of_min_val_in_diff_list]="already matched"

        ## Add the indices (a list) to a bigger list that will hold a list of the N closest values
        ## for each day. 
        #min_indices_list.append(min_indices)

        indices_of_matches.append(min_indices)

    return indices_of_matches


## This function takes a list of lists and another list of lists and groups the first list based
## on the indices in the second one?
def use_list_of_list_of_indices_to_group_a_list_of_lists(main_list,list_of_list_indices):

    main_list_with_criteria=[]

    ## for item in sim days by day (each item will be a list of ints corresponding to days aka list indices)
    for i in range(len(list_of_list_indices)):

        ## an interim variable to hold a single day?
        interim_list=[]

        ## for each list index corresponding to day of year
        for j in range(len(list_of_list_indices[i])):

            ## Go to list with the data you want and collect it
            interim_list.append(main_list[list_of_list_indices[i][j]])

        ## put all the days usages collected into the final list
        main_list_with_criteria.append(interim_list)

        ## the result of this is a list of 366 days, each day has N similar days in it, each similar day has 96 values. 

    return main_list_with_criteria

def zip_all_items_of_a_list(list_to_zip):
    zipped_list=[]
    for item in list_to_zip:
        zipped_list.append(zip(*item))

    return zipped_list

def get_ave_std_of_list_of_list_of_list(list_to_analyze):

    list_average=[]
    list_std=[]
    list_average_plus_stdev=[]
    list_average_minus_stdev=[]

    for list_of_list in list_to_analyze:
        inter_list_ave=[]
        inter_list_std=[]
        inter_list_stdup=[]
        inter_list_stdlo=[]
        for list_ in list_of_list:
            inter=np.array(list_)
            inter_list_ave.append(inter.mean())
            inter_list_std.append(inter.std())
            inter_list_stdup.append(inter.mean()+inter.std())
            inter_list_stdlo.append(inter.mean()-inter.std())
        list_average.append(inter_list_ave)
        list_std.append(inter_list_std)
        list_average_plus_stdev.append(inter_list_stdup)
        list_average_minus_stdev.append(inter_list_stdlo)

    return_list=[list_average,list_average_plus_stdev,list_average_minus_stdev,list_std]

    return return_list

